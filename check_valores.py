from openpyxl import load_workbook

# Carregar com data_only=True para obter valores calculados
wb = load_workbook('3.1_DASH_MENSAL_01_26.xlsx', data_only=True)
ws = wb['Indicadores Usinagem']

print("=== LINHA 28 COM VALORES CALCULADOS ===\n")
print(f"B28: {ws['B28'].value}")
print(f"C28: {ws['C28'].value}")
print(f"D28: {ws['D28'].value}")

print("\n=== CONTEXTO LINHAS 25-28 (O QUE VOCÊ PEDIU) ===\n")
for row in range(25, 29):
    a = ws.cell(row, 1).value
    b = ws.cell(row, 2).value
    c = ws.cell(row, 3).value
    d = ws.cell(row, 4).value
    print(f"Linha {row}: A={a} | B={b} | C={c} | D={d}")

print("\n=== TODAS AS LINHAS 1-30 ===\n")
for row in range(1, 31):
    c = ws.cell(row, 3).value
    d = ws.cell(row, 4).value
    if c is not None or d is not None:
        print(f"Linha {row}: C={c} | D={d}")
