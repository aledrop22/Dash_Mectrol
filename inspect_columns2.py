import pandas as pd
from openpyxl import load_workbook

print("[inspect_columns2] starting")

path = r"c:\Users\xandy\Documents\GitHub\Dash_Mectrol\3.1_DASH_MENSAL_01_26.xlsx"

# pandas default read (first sheet)
try:
    df = pd.read_excel(path, engine='openpyxl', dtype=str).fillna("")
    print("[inspect_columns2] pandas default loaded", df.shape)
    print(df.columns.tolist())
except Exception as e:
    print("[inspect_columns2] pandas error", e)

# inspect all sheets with openpyxl
try:
    wb = load_workbook(path, read_only=True)
    print("[inspect_columns2] sheetnames", wb.sheetnames)
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        print(f"[inspect_columns2] sheet '{sheetname}' first 10 rows:")
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
            print(i, row[:40])
        print("---")
except Exception as e:
    print("[inspect_columns2] openpyxl error", e)

# manually replicate carregar_dados_cronograma logic for this file
try:
    xls = pd.ExcelFile(path, engine='openpyxl')
    colunas_obrigatorias = ['OP', 'Transportadora', 'Previsão', 'Atividade PCP']
    df_choice = None
    chosen_sheet = None
    for sheet in xls.sheet_names:
        try:
            tmp = pd.read_excel(xls, sheet_name=sheet, engine='openpyxl', dtype=str).fillna("")
        except Exception:
            continue
        if all(c in tmp.columns for c in colunas_obrigatorias):
            df_choice = tmp
            chosen_sheet = sheet
            break
    if df_choice is None:
        df_choice = pd.read_excel(path, engine='openpyxl', dtype=str).fillna("")
        chosen_sheet = xls.sheet_names[0]
    print(f"[inspect_columns2] chosen sheet '{chosen_sheet}' df shape", df_choice.shape)
    print(df_choice.columns.tolist())
except Exception as e:
    print("[inspect_columns2] error in manual logic", e)

print("[inspect_columns2] done")
