import pandas as pd
from openpyxl import load_workbook

print("[inspect_columns] starting")

path = r"c:\Users\xandy\Documents\GitHub\Dash_Mectrol\3.1_DASH_MENSAL_01_26.xlsx"

# pandas view
try:
    df = pd.read_excel(path, engine='openpyxl', dtype=str).fillna("")
    print("[inspect_columns] pandas loaded", df.shape)
    print(df.columns.tolist())
except Exception as e:
    print("[inspect_columns] pandas error", e)

# openpyxl raw rows
try:
    wb = load_workbook(path, read_only=True)
    print("[inspect_columns] sheetnames", wb.sheetnames)
    ws = wb.active
    print("[inspect_columns] active sheet", ws.title)
    print("[inspect_columns] first 10 rows (openpyxl):")
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        print(i, row[:20])
except Exception as e:
    print("[inspect_columns] openpyxl error", e)


# agora testar a função de carregamento usada pela app
try:
    from app_qualidade import carregar_dados_cronograma
    df2, msg2 = carregar_dados_cronograma()
    print("[inspect_columns] carregar_dados_cronograma returned", None if df2 is None else df2.shape, msg2)
    if df2 is not None:
        print(df2.columns.tolist())
except Exception as e:
    import traceback
    traceback.print_exc()
    print("[inspect_columns] erro ao chamar carregar_dados_cronograma", e)

print("[inspect_columns] done")
