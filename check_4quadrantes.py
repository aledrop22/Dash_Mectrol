import pandas as pd
import openpyxl

arquivo = '3.1_DASH_MENSAL_01_26.xlsx'
xls = pd.ExcelFile(arquivo)

# Verificar com diferentes headers
print("📋 PRODUTOS DE REFUGO - Verificando estrutura:")
for header_row in [0, 1, 2]:
    try:
        df = pd.read_excel(arquivo, sheet_name='PRODUTOS DE REFUGO', header=header_row, nrows=1)
        print(f"\nCom header={header_row}:")
        print(f"  Total colunas: {len(df.columns)}")
        print("  Primeiras colunas:", df.columns[:5].tolist())
        if len(df.columns) > 20:
            print("  Colunas 0-5:", df.columns[0:6].tolist())
            print("  Colunas 16-23 (Q-W):", df.columns[16:23].tolist() if len(df.columns) > 23 else df.columns[16:].tolist())
            print("  Colunas 23-32 (X-AF):", df.columns[23:32].tolist() if len(df.columns) > 32 else df.columns[23:].tolist())
            print("  Colunas 5-16 (F-P):", df.columns[5:16].tolist())
    except Exception as e:
        print(f"  Erro com header={header_row}: {e}")

# Tentar ler com openpyxl para ver o conteúdo real
print("\n\n📊 Verificando com openpyxl:")
wb = openpyxl.load_workbook(arquivo)
ws = wb['PRODUTOS DE REFUGO']
print(f"Dimensões: {ws.dimensions}")
print("Primeira linha (header):")
for col in range(1, min(15, ws.max_column + 1)):
    print(f"  Col {col}: {ws.cell(1, col).value}")
