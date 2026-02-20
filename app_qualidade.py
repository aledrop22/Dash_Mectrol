import streamlit as st
import pandas as pd
from datetime import date, datetime, timedelta
import os
import re
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

st.set_page_config(page_title="Mectrol Qualidade", layout="wide", page_icon="mectrol.jpg")

# --- GERENCIAMENTO DE PASTAS E DATA ---
hoje = date.today()
mes_ano_atual = hoje.strftime("%m-%y") # Ex: 02-26

# Definição das pastas
PASTA_RNC = f"RNC {mes_ano_atual}"
PASTA_DB = "BANCO_DADOS_MENSAIS"
PASTA_CRONOGRAMA = f"CRONOGRAMA {mes_ano_atual}" 

# Criação automática
for pasta in [PASTA_RNC, PASTA_DB, PASTA_CRONOGRAMA]:
    if not os.path.exists(pasta):
        try: os.makedirs(pasta)
        except: pass

# --- ARQUIVOS ---
NOME_DB_MES = f"Banco_Dados_Qualidade_{mes_ano_atual}.xlsx"
ARQUIVO_HISTORICO = os.path.join(PASTA_DB, NOME_DB_MES)
ARQUIVO_MODELO_RNC = 'USAR MODELO RNC.xlsx' 

COLUNAS_BANCO_PADRAO = [
    'Data Registro', 'Hora Saída', 'Inspetor', 'Inspetor Responsável', 'Data Chegada',
    'Hora Entrada', 'Data Produção', 'OP', 'Cliente', 'Descrição', 'Transportadora',
    'Pedido', 'Codigo', 'Qtd Total', 'Peças produzidas', 'Aprovado', 'Reprovado',
    'Tipo Refugo', 'Sobra Medida 1', 'Sobra Medida 2', 'Maquina', 'Operador',
    'Cód. Castanha', 'Bat. Esq', 'Bat. Dir', 'Emp. Esq', 'Emp. Dir', 'Obs', 'Status',
    'Aprovação Especial', 'RNC Fuso', 'RNC Castanha', 'RNC Guia', 'RNC Bloco',
    'Fuso Retificado', 'Fuso Retificado Adaptado',
    'Castanha Retificada Adaptada', 'Fuso Laminado', 'Fuso Laminado PRECISÃO',
    'Fuso Laminado Adaptado', 'Castanha Laminada Adaptada', 'Guia', 'Bloco',
    'Usinagem', 'Inspeção', 'Desenho', 'Programação CNC', 'Produção', 'Gerar op',
    'PCP', 'Medida não conforme', 'Usinagem não conforme', 'Acabamento Ruim',
    'Concentricidade', 'Craterizada', 'Estética', 'Rebarba', 'Faltou Chaveta',
    'Desenho Errado', 'Código Castanha', 'Motivo_Usinagem', 'Motivo_Medida',
    'Motivo_Outros'
]

# --- SESSÃO ---
if 'pagina' not in st.session_state: st.session_state.pagina = "🏠 Home"
if 'tipo_relogio_key' not in st.session_state: st.session_state.tipo_relogio_key = "Centesimal"
if 'relogio_anterior' not in st.session_state: st.session_state.relogio_anterior = "Centesimal"
if 'df_manual' not in st.session_state: st.session_state.df_manual = None 
if 'peca_atual' not in st.session_state: st.session_state.peca_atual = 1
if 'pecas_inspecionadas' not in st.session_state: st.session_state.pecas_inspecionadas = {}
if 'op_anterior' not in st.session_state: st.session_state.op_anterior = None

for k in ['emp_e', 'bat_e', 'bat_d', 'emp_d']:
    if k not in st.session_state: st.session_state[k] = None

# --- LISTAS PADRÃO ---
LISTA_MAQUINAS = ["", "CNC-01", "CNC-30", "GL-01", "GL-02", "FRESA-01", "FRESA-02", "TORNO -01", "TORNO -02", "TORNO -03"]
LISTA_TORNEIROS = ["", "Everton", "Alex", "Pedro", "Leandro", "Vitor", "Vinicius", "Rodrigo", "Marcos", "Luiz", "Lucas"]

# --- CSS (LAYOUT RESTAURADO) ---
st.markdown("""
    <style>
    .main .block-container { padding-top: 1rem !important; }
    .stMarkdown, .stSelectbox label, .stTextInput label, .stNumberInput label, .stDateInput label, .stTextArea label, .stRadio label { font-size: 16px !important; }
    div.stButton > button:first-child { font-weight: bold; width: 100%; }
    .blink-missing {
        color: #b00020;
        font-weight: 600;
        margin: 4px 0 6px 0;
        animation: blinkFade 1.6s ease-in-out infinite;
    }
    @keyframes blinkFade {
        0% { opacity: 1; }
        50% { opacity: 0.35; }
        100% { opacity: 1; }
    }
    
    /* CARD GRANDE (ESQUERDA) */
    .total-card {
        background-color: #f8f9fa;
        border: 2px solid #dee2e6;
        border-radius: 12px;
        text-align: center;
        padding: 60px 20px;
        height: 100%;
        box-shadow: 0 4px 6px rgba(0,0,0,0.1);
        display: flex;
        flex-direction: column;
        justify-content: center;
    }
    .total-num { font-size: 60px; font-weight: 800; color: #0d6efd; line-height: 1; }
    .total-label { font-size: 20px; color: #6c757d; font-weight: 600; margin-top: 10px; }

    /* CARDS TRANSPORTADORA (DIREITA) */
    .transp-card {
        background-color: #fff;
        border-radius: 8px;
        padding: 15px;
        margin-bottom: 10px;
        box-shadow: 0 2px 4px rgba(0,0,0,0.05);
        border: 1px solid #e0e0e0;
        position: relative;
        overflow: hidden;
        height: 90px;
        display: flex;
        align-items: center;
        justify-content: space-between;
    }
    /* Cores das Bordas */
    .border-alta { border-left: 10px solid #dc3545; }
    .border-media { border-left: 10px solid #ffc107; }
    .border-baixa { border-left: 10px solid #28a745; }
    .border-outra { border-left: 10px solid #6c757d; }

    .t-name { font-weight: 700; font-size: 15px; color: #333; display: block; line-height: 1.2; }
    .t-val { font-weight: 800; font-size: 32px; color: #333; margin-left: 10px; }
    
    /* Botões do Filtro Transp (Escondidos visualmente, mas funcionais sobre os cards se quiser, aqui usaremos botões nativos estilizados ou cards puros) */
    
    .tolerancia-alert { background-color: #fff3cd; color: #856404; padding: 15px; border-left: 6px solid #ffc107; border-radius: 5px; margin: 10px 0; }
    .isento-alert { background-color: #d1ecf1; color: #0c5460; padding: 15px; border-left: 6px solid #17a2b8; border-radius: 5px; margin: 10px 0; }
    </style>
""", unsafe_allow_html=True)

# --- FUNÇÕES ---
def ajustar_casas_relogio():
    """Converte todos os valores de medição quando o tipo de relógio muda."""
    try:
        novo = st.session_state.tipo_relogio_key
        antigo = st.session_state.get('relogio_anterior', "Centesimal")
        
        # Se não mudou, não faz nada
        if novo == antigo:
            return
        
        # Calcula o fator de conversão
        fator = 0.1 if "Milesimal" in novo and "Centesimal" in antigo else 10.0 if "Centesimal" in novo and "Milesimal" in antigo else 1.0
        
        if fator != 1.0:
            # Ajusta chaves de peça única
            for k in ['emp_e', 'bat_e', 'bat_d', 'emp_d']:
                val = st.session_state.get(k)
                if val is not None and val != "":
                    try:
                        # Converter tanto strings quanto números
                        if isinstance(val, str):
                            num_val = float(val)
                        else:
                            num_val = float(val)
                        convertido = round(num_val * fator, 3)
                        st.session_state[k] = str(convertido) if isinstance(val, str) else convertido
                    except Exception as e:
                        pass
            
            # Ajusta chaves de múltiplas peças
            peca_atual = st.session_state.get('peca_atual', 1)
            for k in ['emp_e', 'bat_e', 'bat_d', 'emp_d']:
                key_multiplo = f"{k}_{peca_atual}"
                val = st.session_state.get(key_multiplo)
                if val is not None and val != "":
                    try:
                        # Converter tanto strings quanto números
                        if isinstance(val, str):
                            num_val = float(val)
                        else:
                            num_val = float(val)
                        convertido = round(num_val * fator, 3)
                        st.session_state[key_multiplo] = str(convertido) if isinstance(val, str) else convertido
                    except Exception as e:
                        pass
        
        st.session_state.relogio_anterior = novo
        # Força recarga para atualizar os campos na UI
        st.rerun()
    except Exception as e:
        # Silenciosamente ignora qualquer erro
        pass

def resetar_classificacoes():
    """Reseta os valores de checkbox de classificação quando a OP muda."""
    chaves_checkbox = [
        'fuso_ret', 'fuso_ret_ad', 'cast_ret', 'cast_ret_ad',
        'fuso_lam', 'fuso_lam_ad', 'cast_lam', 'cast_lam_ad',
        'guia', 'bloco'
    ]
    for chave in chaves_checkbox:
        if chave in st.session_state:
            del st.session_state[chave]

def converter_medida(key_name):
    """Converte valor inteiro para decimal conforme o tipo de relógio selecionado."""
    try:
        if key_name in st.session_state and st.session_state[key_name]:
            val_str = str(st.session_state[key_name]).strip()
            if val_str:
                val = float(val_str)
                tipo = st.session_state.get("tipo_relogio_key", "Centesimal")
                
                # Só converte se o valor for >= 1.0 (indicando que foi digitado um número inteiro)
                if val >= 1.0:
                    divisor = 1000.0 if "Milesimal" in tipo else 100.0
                    st.session_state[key_name] = str(round(val / divisor, 3))
    except Exception as e:
        # Silenciosamente ignora erros de conversão
        pass

def formatar_hora_automatica(key_hora=None):
    """Formata automaticamente entrada de hora (753 -> 07:53, 15 -> 15:00, etc)"""
    # Se não especificou a key, tentar detectar
    if key_hora is None:
        if "hora_entrada_key" in st.session_state:
            key_hora = "hora_entrada_key"
        else:
            # Procurar keys de hora para multi-peça
            for key in st.session_state.keys():
                if isinstance(key, str) and key.startswith("hora_entrada_") and key != "hora_entrada_key":
                    key_hora = key
                    break
    
    if key_hora and key_hora in st.session_state:
        val = st.session_state[key_hora]
        limpo = ''.join(filter(str.isdigit, str(val)))
        if len(limpo) in [1, 2]:
            try:
                h = int(limpo)
                if 0 <= h <= 23: st.session_state[key_hora] = f"{h:02d}:00 - {'Manhã' if h < 12 else 'Tarde' if h < 18 else 'Noite'}"
            except: pass
        elif len(limpo) == 3: limpo = "0" + limpo
        if len(limpo) == 4:
            try:
                h = int(limpo[:2]); m = int(limpo[2:])
                if 0 <= h <= 23 and 0 <= m <= 59: st.session_state[key_hora] = f"{h:02d}:{m:02d} - {'Manhã' if h < 12 else 'Tarde' if h < 18 else 'Noite'}"
            except: pass

def definir_prioridade(transportadora):
    transp = str(transportadora).upper()
    if "MIGUEL" in transp: return "1_ALTA"
    elif "BIGTRANS" in transp: return "2_MEDIA"
    elif "ALFA" in transp or "RODONAVES" in transp: return "3_BAIXA"
    else: return "4_OUTRAS"

def arredondar_sobra_10(valor):
    try: return int(float(str(valor).replace(',', '.')) // 10) * 10
    except: return 0

def extrair_medidas_peca(descricao):
    try:
        match = re.search(r'-([\d.,]+)-([\d.,]+)-', descricao)
        if match: return arredondar_sobra_10(match.group(1).replace('.', '').replace(',', '.'))
        return 0
    except: return 0

def verificar_isencao_tamanho(descricao):
    try:
        desc_upper = descricao.upper()
        if "FUSO" not in desc_upper: return False, False
        match_diam = re.search(r'R(\d+)', desc_upper)
        diametro = int(match_diam.group(1)) if match_diam else 0
        match_len = re.search(r'-([\d.,]+)-0[,.]0', desc_upper)
        comp = float(match_len.group(1).replace('.', '').replace(',', '.')) if match_len else 0.0
        return (True, True) if diametro in [63, 80, 100] and comp > 2400 else (True, False)
    except: return False, False

def detectar_tipo_componente(descricao):
    """Detecta se é FUSO, GUIA ou BLOCO baseado em padrões específicos."""
    desc_upper = descricao.upper()
    
    # GUIAS: Iniciais HG, RG, EG, WER, MGN, MGW (buscar em toda a descrição)
    padroes_guias = [r'HG', r'RG', r'EG', r'WER', r'MGN', r'MGW']
    eh_guia = any(padrao in desc_upper for padrao in padroes_guias)
    
    # FUSOS: Iniciais com R ou L, seguidas de números OU palavra "FUSO" na descrição
    padroes_fusos = [r'\b[2485]R\b', r'\bR\d+\b', r'\bL\d+\b', r'FUSO']
    eh_fuso = any(re.search(p, desc_upper) if p in [r'\b[2485]R\b', r'\bR\d+\b', r'\bL\d+\b'] else p in desc_upper for p in padroes_fusos)
    
    # BLOCO: Palavras-chave específicas
    eh_bloco = 'BLOCO' in desc_upper or 'PATIM' in desc_upper
    
    return eh_fuso, eh_guia, eh_bloco

def detectar_classe_precisao(descricao):
    """Detecta classe de precisão (tolerância) na descrição e retorna classificação automática."""
    desc_upper = descricao.upper()
    
    # 1. DETECTAR TIPO DE COMPONENTE
    eh_fuso, eh_guia, eh_bloco = detectar_tipo_componente(descricao)
    
    # 2. EXTRAIR TOLERÂNCIA (0,025 ou 0.025 ou 0,05 ou 0.05, etc)
    match_tolerancia = re.search(r'[-,.]?(0[,/.](0\d{2}|0\d{3}|\d{2}|\d{3}))', desc_upper)
    tolerancia_encontrada = None
    
    if match_tolerancia:
        tol_str = match_tolerancia.group(1).replace(',', '.').replace('/', '.')
        try:
            tolerancia_encontrada = float(tol_str)
        except:
            pass
    
    # 3. VERIFICAR CARACTERÍSTICAS DE PROCESSAMENTO
    tem_mop = 'MOP' in desc_upper
    tem_esferas = 'ESFERAS' in desc_upper
    eh_conjunto = 'CONJUNTO' in desc_upper
    eh_retificado = 'RETIFICADO' in desc_upper
    eh_laminado = 'LAMINADO' in desc_upper
    eh_adaptado = 'ADAPTADO' in desc_upper
    eh_castanha = 'CASTANHA' in desc_upper or 'ADAPTADA' in desc_upper
    
    # 4. INICIALIZAR RESULTADO
    resultado = {
        'sel_fuso_ret': False,
        'sel_fuso_ret_ad': False,
        'sel_cast_ret': False,
        'sel_cast_ret_ad': False,
        'sel_fuso_lam': False,
        'sel_fuso_lam_ad': False,
        'sel_cast_lam': False,
        'sel_cast_lam_ad': False,
        'sel_guia': False,
        'sel_bloco': False,
    }
    
    # 5. CLASSIFICAR COMPONENTES ESPECIAIS
    if eh_guia:
        resultado['sel_guia'] = True
        return resultado
    
    if eh_bloco:
        resultado['sel_bloco'] = True
        return resultado
    
    # 6. CLASSIFICAR FUSOS E CASTANHAS
    if eh_fuso or eh_castanha:
        # Tipo de componente para escolher a chave correta
        tipo = 'fuso' if eh_fuso else 'cast'
        
        # PRIORIDADE 1: Palavras explícitas RETIFICADO ou LAMINADO têm prioridade MÁXIMA
        if eh_retificado and not eh_laminado:
            chave_ret = f'sel_{tipo}_ret'
            chave_ret_ad = f'sel_{tipo}_ret_ad'
            resultado[chave_ret_ad if eh_adaptado else chave_ret] = True
            return resultado
        
        if eh_laminado and not eh_retificado:
            chave_lam = f'sel_{tipo}_lam'
            chave_lam_ad = f'sel_{tipo}_lam_ad'
            resultado[chave_lam_ad if eh_adaptado else chave_lam] = True
            return resultado
        
        # PRIORIDADE 2: MOP → RETIFICADO
        if tem_mop:
            chave_ret = f'sel_{tipo}_ret'
            chave_ret_ad = f'sel_{tipo}_ret_ad'
            resultado[chave_ret_ad if eh_adaptado else chave_ret] = True
            return resultado
        
        # PRIORIDADE 3: LAMINADO DE PRECISÃO
        # (ESFERAS + CONJUNTO + tolerância baixa (0.023 ou 0.05))
        if tem_esferas and eh_conjunto and tolerancia_encontrada and tolerancia_encontrada in [0.023, 0.05]:
            chave_lam = f'sel_{tipo}_lam'
            chave_lam_ad = f'sel_{tipo}_lam_ad'
            resultado[chave_lam_ad if eh_adaptado else chave_lam] = True
            return resultado
        
        # PRIORIDADE 4: Inferir por tolerância
        # Tolerâncias baixas (<= 0.023) indicam RETIFICADO
        # Tolerâncias altas (>= 0.05) indicam LAMINADO
        if tolerancia_encontrada:
            if tolerancia_encontrada <= 0.023:
                chave_ret = f'sel_{tipo}_ret'
                chave_ret_ad = f'sel_{tipo}_ret_ad'
                resultado[chave_ret_ad if eh_adaptado else chave_ret] = True
            elif tolerancia_encontrada >= 0.05:
                chave_lam = f'sel_{tipo}_lam'
                chave_lam_ad = f'sel_{tipo}_lam_ad'
                resultado[chave_lam_ad if eh_adaptado else chave_lam] = True
            return resultado
    
    return resultado

def tratar_valor_numerico_string(valor):
    s = str(valor)
    return s[:-2] if s.endswith('.0') else s

def gerar_nome_arquivo_rnc(pedido, cliente, item):
    limpo = re.sub(r'[\\/*?:"<>|]', "", f"RNC - {tratar_valor_numerico_string(pedido)} - {str(cliente).strip().split(' ')[0]} - {tratar_valor_numerico_string(item)}.xlsx")
    return os.path.join(PASTA_RNC, limpo)

# --- BUSCA SUPER FLEXÍVEL DO CRONOGRAMA ---
def buscar_arquivo_cronograma(pasta_busca):
    try:
        if not os.path.exists(pasta_busca): return None, f"A pasta '{pasta_busca}' não existe."
        arquivos = os.listdir(pasta_busca)
        
        # Procura qualquer Excel que tenha CRONOGRAMA no nome (Maiusculo ou minusculo)
        candidatos = [f for f in arquivos if "CRONOGRAMA" in f.upper() and f.endswith(".xlsx") and not f.startswith("~$")]
        
        if not candidatos:
            return None, f"Pasta '{pasta_busca}' encontrada, mas vazia ou sem Excel."
        
        candidatos.sort(reverse=True) # Pega o mais recente
        return os.path.join(pasta_busca, candidatos[0]), candidatos[0]
    except Exception as e: return None, str(e)


# helpers for RNC generation --------------------------------------------------

def _campos_para_rnc(dados_rnc: dict) -> list:
    """Return a list of checkbox labels that should be marked on the RNC template.

    The template already contains checkboxes for **RETRABALHO**, **REFUGO**,
    **SOBRA** and **ANALISE**; this helper adds the three "motivo" flags when
    the operator marcou um retrabalho (usinagem/medida/outros).  Separar esta
    lógica facilita testes unitários e evita concatenar o motivo como texto
    (o usuário pediu "sem previsão no texto").
    """
    campos = []
    tipo = str(dados_rnc.get('Tipo_Refugo', '')).upper()
    if tipo == "RETRABALHO":
        campos.append("RETRABALHO")
    if "MORTE" in tipo:
        campos.append("REFUGO")
    if "SOBRA" in tipo:
        campos.append("SOBRA")
    if dados_rnc.get('Analise'):
        campos.append("ANALISE")

    # motivos específicos apresentados com checkboxes no formulário
    if dados_rnc.get('Motivo_Usinagem'):
        campos.append("USINAGEM")
    if dados_rnc.get('Motivo_Medida'):
        campos.append("MEDIDA")
    if dados_rnc.get('Motivo_Outros'):
        campos.append("OUTROS")
    return campos


def preencher_modelo_rnc_existente(dados_rnc, nome_saida):
    try:
        if not os.path.exists(ARQUIVO_MODELO_RNC):
            return False
        wb = load_workbook(ARQUIVO_MODELO_RNC)
        ws = wb.active
        ws.row_dimensions[6].height = 21

        mapa = {
            "Data:": date.today().strftime("%d/%m/%Y"),
            "OP": tratar_valor_numerico_string(dados_rnc['OP']),
            "Item": dados_rnc['Descricao'],
            "Quantidade na OP": dados_rnc['Qtd_Total'],
            "Quantidade de Falha": dados_rnc['Qtd_Reprovada'],
            "PV": tratar_valor_numerico_string(dados_rnc.get('Pedido', '')),
            "Cliente": dados_rnc['Cliente'],
            "Máquina": dados_rnc['Maquina'],
            "Operador": dados_rnc['Operador'],
            # a descrição do ocorrido deixa de ter os motivos, pois eles são
            # representados por checkboxes separados no template
            "Descrição do ocorrido": dados_rnc.get('Descricao_Ocorrido', ''),
            "Observação do colaborador": dados_rnc.get('Obs_Colaborador', ''),
            "Possível causa": dados_rnc.get('Obs_Inspetor', ''),
            "Medida Sobra (se houver)": f"{dados_rnc.get('Sobra1')} mm" if dados_rnc.get('Sobra1') else "",
            "Responsável": dados_rnc.get('Inspetor', '')
        }

        campos_x = _campos_para_rnc(dados_rnc)

        def get_master(sheet, r, c):
            for m in sheet.merged_cells.ranges:
                if r >= m.min_row and r <= m.max_row and c >= m.min_col and c <= m.max_col:
                    return sheet.cell(m.min_row, m.min_col)
            return sheet.cell(r, c)

        for row in ws.iter_rows(min_row=1, max_row=60, min_col=1, max_col=10):
            for cell in row:
                if cell.value:
                    val_str = str(cell.value).strip()
                    for k, v in mapa.items():
                        if val_str.startswith(k):
                            tgt = get_master(ws, cell.row, cell.column + 1)
                            tgt.value = v
                            tgt.alignment = Alignment(wrap_text=True, vertical='center',
                                                       horizontal='center' if k == "Responsável" else 'left')
                    if val_str in ["RETRABALHO", "REFUGO", "SOBRA", "ANALISE", "USINAGEM", "MEDIDA", "OUTROS"]:
                        tgt = get_master(ws, cell.row, cell.column + 1)
                        tgt.value = "X" if val_str in campos_x else ""
                        if tgt.value == "X":
                            tgt.font = Font(bold=True, size=12, name='Arial')
                            tgt.alignment = Alignment(horizontal='center', vertical='center')

        wb.save(os.path.join(PASTA_RNC, nome_saida))
        return True
    except:
        return False

@st.cache_data(ttl=600)
def carregar_dados_cronograma():
    """Carrega cronograma com validação robusta de colunas necessárias."""
    path, nome = buscar_arquivo_cronograma(PASTA_CRONOGRAMA)
    if not path: 
        return None, nome
    try:
        # Ler todas as sheets e escolher a que contém as colunas necessárias.
        # alguns cronogramas (como o 3.1_DASH_MENSAL_01_26.xlsx) têm primeira
        # aba apenas com textos e a aba real chama-se 'Lançamentos'.
        xls = pd.ExcelFile(path, engine='openpyxl')
        # There are different cronograma formats.  Older files contain Transportadora
        # and Previsão, while the monthly "DASH" files simply have OP/Cliente/etc.
        colunas_obrigatorias = ['OP', 'Transportadora', 'Previsão', 'Atividade PCP']
        df = None

        # helper to determine if a sheet looks like a cronograma table
        def sheet_looks_like_data(tmp):
            # must have at least two rows and two columns
            if tmp.shape[0] < 2 or tmp.shape[1] < 2:
                return False
            if 'OP' not in tmp.columns:
                return False
            return True

        # prefer a sheet named like "lanç" (lanÃ§amentos) or containing the
        # expected required columns
        for sheet in xls.sheet_names:
            try:
                tmp = pd.read_excel(xls, sheet_name=sheet, engine='openpyxl', dtype=str).fillna("")
            except Exception:
                continue
            if not sheet_looks_like_data(tmp):
                continue
            # sheet name hint is strongest
            if "lan" in sheet.lower():
                df = tmp
                break
            # otherwise fall back to the original required-column check
            if all(c in tmp.columns for c in colunas_obrigatorias):
                df = tmp
                break

        # last resort: pick first sheet that at least has an OP column
        if df is None:
            for sheet in xls.sheet_names:
                try:
                    tmp = pd.read_excel(xls, sheet_name=sheet, engine='openpyxl', dtype=str).fillna("")
                except Exception:
                    continue
                if sheet_looks_like_data(tmp):
                    df = tmp
                    break

        if df is None:
            # última tentativa: ler a primeira planilha do arquivo
            df = pd.read_excel(path, engine='openpyxl', dtype=str).fillna("")

        # Validação de colunas obrigatórias
        colunas_faltantes = [c for c in colunas_obrigatorias if c not in df.columns]
        if colunas_faltantes:
            return None, f"Colunas faltantes: {', '.join(colunas_faltantes)}"

        # Conversão de tipos com conversão segura
        df['OP'] = df['OP'].str.strip()  # Manter como string com zeros à esquerda
        df['Transportadora'] = df['Transportadora'].astype(str).str.strip()
        df['Atividade PCP'] = df['Atividade PCP'].astype(str).str.upper().str.strip()
        df = df[df['OP'] != "0"]  # Remove OP inválidos

        # Preservar colunas numéricas de ID como string (Pedido, Código Item, etc.)
        for col in ['Pedido', 'Código Item', 'Item', 'FT', 'GEF', 'PC']:
            if col in df.columns:
                df[col] = df[col].astype(str).str.strip()

        df['Previsão'] = pd.to_datetime(df['Previsão'], errors='coerce')
        df['Prioridade_Logistica'] = df['Transportadora'].apply(definir_prioridade)
        return df, nome
    except (FileNotFoundError, ValueError, KeyError) as e:
        return None, f"Erro ao ler cronograma: {str(e)}"

def carregar_dados_historico():
    """Carrega histórico de inspeção com tratamento de erros."""
    if os.path.exists(ARQUIVO_HISTORICO):
        try:
            try:
                df_conf = pd.read_excel(ARQUIVO_HISTORICO, sheet_name="CONFORME")
            except Exception:
                df_conf = pd.DataFrame()
            try:
                df_nao = pd.read_excel(ARQUIVO_HISTORICO, sheet_name="NAO CONFORME")
            except Exception:
                df_nao = pd.DataFrame()
            if not df_conf.empty or not df_nao.empty:
                df = pd.concat([df_conf, df_nao], ignore_index=True)
            else:
                df = pd.read_excel(ARQUIVO_HISTORICO)
            if 'Data Registro' in df.columns:
                df['Data Registro'] = pd.to_datetime(df['Data Registro'], format='%d/%m/%Y', errors='coerce')
            return df
        except (FileNotFoundError, ValueError, KeyError):
            return pd.DataFrame()
    return pd.DataFrame()

def obter_colunas_banco():
    """Retorna as colunas oficiais do banco mensal."""
    if os.path.exists(ARQUIVO_HISTORICO):
        try:
            return pd.read_excel(ARQUIVO_HISTORICO, nrows=0).columns.tolist()
        except Exception:
            pass
    return COLUNAS_BANCO_PADRAO

def normalizar_df_banco(df):
    """Garante todas as colunas e preenche vazios com string vazia."""
    cols_base = obter_colunas_banco()
    extras = [c for c in df.columns if c not in cols_base]
    colunas = cols_base + extras
    for col in cols_base:
        if col not in df.columns:
            df[col] = ""
    df = df[colunas]
    return df.fillna("")

def normalizar_arquivo_banco():
    """Normaliza o arquivo do banco mensal existente."""
    try:
        if os.path.exists(ARQUIVO_HISTORICO):
            try:
                df_conf = pd.read_excel(ARQUIVO_HISTORICO, sheet_name="CONFORME")
            except Exception:
                df_conf = pd.DataFrame()
            try:
                df_nao = pd.read_excel(ARQUIVO_HISTORICO, sheet_name="NAO CONFORME")
            except Exception:
                df_nao = pd.DataFrame()
            if df_conf.empty and df_nao.empty:
                df_conf = pd.read_excel(ARQUIVO_HISTORICO)
        else:
            df_conf = pd.DataFrame(columns=COLUNAS_BANCO_PADRAO)
            df_nao = pd.DataFrame(columns=COLUNAS_BANCO_PADRAO)

        df_conf = normalizar_df_banco(df_conf)
        df_nao = normalizar_df_banco(df_nao)
        with pd.ExcelWriter(ARQUIVO_HISTORICO, engine='openpyxl') as writer:
            df_conf.to_excel(writer, sheet_name="CONFORME", index=False)
            df_nao.to_excel(writer, sheet_name="NAO CONFORME", index=False)
    except Exception:
        pass

def salvar_no_banco(df_novo, nao_conforme=False):
    """Salva no banco mensal, separando em abas CONFORME e NAO CONFORME."""
    df_novo = normalizar_df_banco(df_novo)
    if os.path.exists(ARQUIVO_HISTORICO):
        try:
            try:
                df_conf = pd.read_excel(ARQUIVO_HISTORICO, sheet_name="CONFORME")
            except Exception:
                df_conf = pd.DataFrame()
            try:
                df_nao = pd.read_excel(ARQUIVO_HISTORICO, sheet_name="NAO CONFORME")
            except Exception:
                df_nao = pd.DataFrame()
            if df_conf.empty and df_nao.empty:
                df_conf = pd.read_excel(ARQUIVO_HISTORICO)
            df_conf = normalizar_df_banco(df_conf)
            df_nao = normalizar_df_banco(df_nao)
        except Exception:
            df_conf = pd.DataFrame()
            df_nao = pd.DataFrame()
    else:
        df_conf = pd.DataFrame()
        df_nao = pd.DataFrame()

    if nao_conforme:
        df_final_nao = pd.concat([df_nao, df_novo], ignore_index=True)
        df_final_conf = df_conf
    else:
        df_final_conf = pd.concat([df_conf, df_novo], ignore_index=True)
        df_final_nao = df_nao

    df_final_conf = normalizar_df_banco(df_final_conf)
    df_final_nao = normalizar_df_banco(df_final_nao)

    with pd.ExcelWriter(ARQUIVO_HISTORICO, engine='openpyxl') as writer:
        df_final_conf.to_excel(writer, sheet_name="CONFORME", index=False)
        df_final_nao.to_excel(writer, sheet_name="NAO CONFORME", index=False)

def aviso_campo_faltante(chave, texto):
    """Exibe um aviso visual leve para campos obrigatorios ausentes."""
    if chave in st.session_state.get('campos_faltantes', set()):
        st.markdown(f"<div class='blink-missing'>⚠ {texto}</div>", unsafe_allow_html=True)

def carregar_dados_refugo():
    """Carrega dados de refugo da aba PRODUTOS DE REFUGO do cronograma."""
    path, nome = buscar_arquivo_cronograma(PASTA_CRONOGRAMA)
    if not path:
        return None
    try:
        # Ler aba PRODUTOS DE REFUGO com header na linha 2 (index 1)
        df = pd.read_excel(path, sheet_name='PRODUTOS DE REFUGO', 
                          header=1, engine='openpyxl', dtype=str)
        df = df.fillna("")
        return df
    except Exception:
        return None

def obter_quadrante_1_motivos():
    """Extrai 1º quadrante: Colunas R-AB da aba Lançamentos (Motivos)."""
    path, _ = buscar_arquivo_cronograma(PASTA_CRONOGRAMA)
    if not path:
        return {"Usinagem": False, "Inspeção": False, "Desenho": False,
                "Programação CNC": False, "Produção": False, "Comercial": False,
                "PCP": False, "RETRABALHO OUTROS DP": False}
    try:
        df = pd.read_excel(path, sheet_name='Lançamentos', nrows=0)
        cols = df.columns[17:28].tolist() if len(df.columns) >= 28 else []
        return {col: False for col in cols} if cols else {"Usinagem": False, "Inspeção": False}
    except:
        return {"Usinagem": False, "Inspeção": False}

def obter_quadrante_2_colunasQW():
    """Extrai 2º quadrante: Colunas Q-W da aba PRODUTOS DE REFUGO."""
    path, _ = buscar_arquivo_cronograma(PASTA_CRONOGRAMA)
    if not path:
        return {}
    try:
        df = pd.read_excel(path, sheet_name='PRODUTOS DE REFUGO', header=1, nrows=1)
        cols = df.columns[16:23] if len(df.columns) >= 23 else []
        return {col: "" for col in cols}
    except:
        return {}

def obter_quadrante_3_colunasXAF():
    """Extrai 3º quadrante: Colunas X-AF da aba PRODUTOS DE REFUGO."""
    path, _ = buscar_arquivo_cronograma(PASTA_CRONOGRAMA)
    if not path:
        return {}
    try:
        df = pd.read_excel(path, sheet_name='PRODUTOS DE REFUGO', header=1, nrows=1)
        cols = df.columns[23:32] if len(df.columns) >= 32 else []
        return {col: "" for col in cols}
    except:
        return {}

def obter_quadrante_4_colunasFP():
    """Extrai 4º quadrante: Colunas F-P da aba PRODUTOS DE REFUGO."""
    path, _ = buscar_arquivo_cronograma(PASTA_CRONOGRAMA)
    if not path:
        return {}
    try:
        df = pd.read_excel(path, sheet_name='PRODUTOS DE REFUGO', header=1, nrows=1)
        cols = df.columns[5:16] if len(df.columns) >= 16 else []
        return {col: "" for col in cols}
    except:
        return {}

if 'banco_normalizado' not in st.session_state:
    normalizar_arquivo_banco()
    st.session_state.banco_normalizado = True

# ==========================================
# INTERFACE - NAVEGAÇÃO NO TOPO
# ==========================================

nav_col1, nav_col2, nav_col3, nav_col4, nav_col5 = st.columns(5)
if nav_col1.button("🏠 Home", use_container_width=True, type="secondary"):
    st.session_state.pagina = "🏠 Home"
if nav_col2.button("🔍 Inspeção", use_container_width=True, type="secondary"):
    st.session_state.pagina = "🔍 Inspeção"
if nav_col3.button("📦 Pré Carga", use_container_width=True, type="secondary"):
    st.session_state.pagina = "📦 Pré Carga"
if nav_col4.button("♻️ Análise Refugo", use_container_width=True, type="secondary"):
    st.session_state.pagina = "♻️ Análise Refugo"
if nav_col5.button("📊 Indicadores", use_container_width=True, type="secondary"):
    st.session_state.pagina = "📊 Indicadores"

st.markdown("---")
pagina = st.session_state.pagina

# SIDEBAR - FILTROS
# ==========================================
with st.sidebar:
    st.image("mectrol.jpg", width=220)
    st.markdown("### Filtros")
    st.markdown("---")
    
    st.markdown("**📅 Data e Período**")
    data_filtro = st.date_input("Selecione o Dia:", date.today(), format="DD/MM/YYYY")
    
    ver_todas_datas = st.checkbox("Ver Todas as Datas", value=False)
    
    usar_periodo = False
    if not ver_todas_datas:
        data_ini, data_fim = data_filtro, data_filtro
        
        usar_periodo = st.checkbox("Análise por Período")
        if usar_periodo:
            col_p1, col_p2 = st.columns(2)
            data_ini = col_p1.date_input("Início", date.today() - timedelta(days=7), format="DD/MM/YYYY")
            data_fim = col_p2.date_input("Fim", date.today(), format="DD/MM/YYYY")
    else:
        data_ini, data_fim = None, None

    st.markdown("---")
    st.markdown("**Prioridade Logística**")
    filtro_prio = []
    if st.checkbox("🔴 1º Expresso São Miguel - Alta", value=True): filtro_prio.append("1_ALTA")
    if st.checkbox("🟡 2º New BigTrans - Média", value=True): filtro_prio.append("2_MEDIA")
    if st.checkbox("🟢 3º Alfa e Rodonaves - Baixa", value=True): filtro_prio.append("3_BAIXA")
    if st.checkbox("⚪ Outras", value=True): filtro_prio.append("4_OUTRAS")
    
    st.markdown("---")
    LISTA_ATIV = ["U - USINAGEM", "I - Industrialização", "H - Indust/Usin", "P - Producao", "S - Separacao"]
    filtro_ativ = [a for a in LISTA_ATIV if st.checkbox(a, value=("USINAGEM" in a))]

# CARGA DADOS (Automático ou Manual)
df_crono = None
nome_crono_display = ""
msg_erro_load = ""

# 1. Tenta carregar automático
if st.session_state.df_manual is None:
    df_auto, info_msg = carregar_dados_cronograma()
    if df_auto is not None and not df_auto.empty:
        df_crono = df_auto
        nome_crono_display = info_msg
    else:
        msg_erro_load = info_msg
else:
    df_crono = st.session_state.df_manual
    nome_crono_display = "Arquivo Carregado Manualmente"

df_hist = carregar_dados_historico()

# FILTRAGEM
df_final = pd.DataFrame()
if df_crono is not None and not df_crono.empty:
    df_filtrado = df_crono.copy()
    if not ver_todas_datas:
        df_filtrado = df_filtrado[
            (df_filtrado['Previsão'].dt.date >= data_ini) & 
            (df_filtrado['Previsão'].dt.date <= data_fim)
        ]
    
    if not df_filtrado.empty:
        df_final = df_filtrado[
            (df_filtrado['Atividade PCP'].isin(filtro_ativ)) & 
            (df_filtrado['Prioridade_Logistica'].isin(filtro_prio))
        ].copy()
        
        # Adicionar coluna de Prazo de Entrega (dias até a previsão)
        df_final['Prazo Entrega (dias)'] = (df_final['Previsão'] - pd.Timestamp(date.today())).dt.days

# CALCULO KPI
qtd_concluidas = 0
delta_txt = "Sem dados ant."
if not df_hist.empty and not ver_todas_datas:
    mask_atual = (df_hist['Data Registro'].dt.date >= data_ini) & (df_hist['Data Registro'].dt.date <= data_fim)
    qtd_concluidas = len(df_hist[mask_atual])
    
    dias_diff = (data_fim - data_ini).days + 1
    ini_ant = data_ini - timedelta(days=dias_diff)
    fim_ant = data_ini - timedelta(days=1)
    mask_ant = (df_hist['Data Registro'].dt.date >= ini_ant) & (df_hist['Data Registro'].dt.date <= fim_ant)
    qtd_ant = len(df_hist[mask_ant])
    
    diff = qtd_concluidas - qtd_ant
    perc = (diff / qtd_ant * 100) if qtd_ant > 0 else 0
    delta_txt = f"{diff:+} ({perc:.0f}%)" if qtd_ant > 0 else f"{diff:+}"

# ==========================================
# PÁGINA: HOME
# ==========================================
if pagina == "🏠 Home":
    c_tit, c_btn = st.columns([0.85, 0.15])
    with c_tit: st.title("📊 Painel de Visão Geral")
    with c_btn:
        if st.button("🔄 Atualizar"): st.cache_data.clear(); st.rerun()

    # --- SEGURANÇA: MENSAGEM SE NÃO CARREGAR ---
    if df_crono is None:
        st.error(f"🚨 Não carregou.")
        with st.expander("🔍 Detalhes e Upload Manual", expanded=True):
            st.write(f"Status: {msg_erro_load}")
            st.info(f"O sistema procurou em: **{PASTA_CRONOGRAMA}**")
            st.warning("Se o arquivo não estiver lá, faça upload abaixo:")
            uploaded_file = st.file_uploader("Arraste o Cronograma aqui", type=["xlsx"])
            if uploaded_file:
                try:
                    # Use the same smart sheet-selection logic as carregar_dados_cronograma
                    xls = pd.ExcelFile(uploaded_file, engine='openpyxl')
                    colunas_obrigatorias = ['OP', 'Transportadora', 'Previsão', 'Atividade PCP']
                    df_up = None

                    # helper to determine if a sheet looks like a cronograma table
                    def sheet_looks_like_data(tmp):
                        if tmp.shape[0] < 2 or tmp.shape[1] < 2:
                            return False
                        if 'OP' not in tmp.columns:
                            return False
                        return True

                    # prefer a sheet named like "lançamentos" or containing required columns
                    for sheet in xls.sheet_names:
                        try:
                            tmp = pd.read_excel(xls, sheet_name=sheet, engine='openpyxl', dtype=str).fillna("")
                        except Exception:
                            continue
                        if not sheet_looks_like_data(tmp):
                            continue
                        if "lan" in sheet.lower():
                            df_up = tmp
                            break
                        if all(c in tmp.columns for c in colunas_obrigatorias):
                            df_up = tmp
                            break

                    # last resort: pick first sheet that at least has an OP column
                    if df_up is None:
                        for sheet in xls.sheet_names:
                            try:
                                tmp = pd.read_excel(xls, sheet_name=sheet, engine='openpyxl', dtype=str).fillna("")
                            except Exception:
                                continue
                            if sheet_looks_like_data(tmp):
                                df_up = tmp
                                break

                    if df_up is None:
                        df_up = pd.read_excel(uploaded_file, engine='openpyxl', dtype=str).fillna("")

                    df_up['OP'] = df_up['OP'].astype(str).str.strip()
                    df_up = df_up[df_up['OP'] != "0"]
                    
                    # Handle Transportadora column if it exists (for compatibility)
                    if 'Transportadora' in df_up.columns:
                        df_up['Transportadora'] = df_up['Transportadora'].astype(str).str.strip()
                        df_up['Prioridade_Logistica'] = df_up['Transportadora'].apply(definir_prioridade)
                    else:
                        df_up['Prioridade_Logistica'] = '4_OUTRAS'  # default
                    
                    # Ensure Previsão is a datetime
                    if 'Previsão' in df_up.columns:
                        df_up['Previsão'] = pd.to_datetime(df_up['Previsão'], errors='coerce')
                    
                    st.session_state.df_manual = df_up
                    st.success("Carregado! Atualizando..."); st.rerun()
                except Exception as e:
                    st.error(f"Erro: {e}")
    
    elif df_crono is not None:
        if not ver_todas_datas:
            st.caption(f"📅 Visualizando: **{data_ini.strftime('%d/%m/%Y')}** até **{data_fim.strftime('%d/%m/%Y')}** | Base: {nome_crono_display}")
        else:
            st.caption(f"📅 Visualizando: **Todo o Cronograma** | Base: {nome_crono_display}")
        
        # Inicializar placeholder para filtro
        if 'filtro_selecionado' not in st.session_state:
            st.session_state.filtro_selecionado = None
        
        # Aplicar filtro aos dados
        df_kpi = df_final.copy()
        if st.session_state.filtro_selecionado:
            df_kpi = df_kpi[df_kpi['Prioridade_Logistica'] == st.session_state.filtro_selecionado]
        
        total_prog = len(df_kpi)
        faltam = max(0, total_prog - qtd_concluidas)

        # --- 1. CARDS DE STATUS (TOPO) ---
        st.markdown("### 📊 Resumo")
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.markdown(f"""
            <div class='transp-card' style='border-left: 8px solid #dc3545;'>
                <span class='t-name'>📦 Total Programado</span>
                <span class='t-val'>{total_prog}</span>
            </div>
            """, unsafe_allow_html=True)
        
        with col2:
            st.markdown(f"""
            <div class='transp-card' style='border-left: 8px solid #28a745;'>
                <span class='t-name'>✅ Peças Concluídas</span>
                <span class='t-val'>{qtd_concluidas}</span>
            </div>
            """, unsafe_allow_html=True)
        
        with col3:
            st.markdown(f"""
            <div class='transp-card' style='border-left: 8px solid #ffc107;'>
                <span class='t-name'>⏳ Faltam Concluir</span>
                <span class='t-val'>{faltam}</span>
            </div>
            """, unsafe_allow_html=True)

        st.markdown("---")

        # --- 2. CARDS DE FILTRO (MEIO) ---
        st.markdown("### 🚚 Filtrar por Transportadora")
        
        counts_all = df_final['Prioridade_Logistica'].value_counts()
        
        map_btn = {
            "1_ALTA": {"label": "1º Expresso São Miguel", "cor": "border-alta"},
            "2_MEDIA": {"label": "2º New BigTrans", "cor": "border-media"},
            "3_BAIXA": {"label": "3º ALFA/RODONAVES", "cor": "border-baixa"},
            "4_OUTRAS": {"label": "4º Outras", "cor": "border-outra"}
        }
        
        col_total_filtro, col_cards_filtro = st.columns([1, 2.5])
        
        with col_total_filtro:
            st.markdown(f"""
                <div class='total-card'>
                    <div class='total-num'>{len(df_final)}</div>
                    <div class='total-label'>TOTAL</div>
                </div>
            """, unsafe_allow_html=True)
        
        with col_cards_filtro:
            c1, c2 = st.columns(2); c3, c4 = st.columns(2)
            
            def render_card_clickable(col, key_prio):
                data = map_btn[key_prio]
                qtd = counts_all.get(key_prio, 0)
                with col:
                    st.markdown(f"""
                    <div class='transp-card {data['cor']}' style='cursor: pointer;'>
                        <span class='t-name'>{data['label']}</span>
                        <span class='t-val'>{qtd}</span>
                    </div>
                    """, unsafe_allow_html=True)
                    
                    if st.button(f"Selecionar", key=f"btn_filtro_{key_prio}", use_container_width=True):
                        st.session_state.filtro_selecionado = key_prio if st.session_state.filtro_selecionado != key_prio else None
                        st.rerun()

            render_card_clickable(c1, "1_ALTA")
            render_card_clickable(c2, "2_MEDIA")
            render_card_clickable(c3, "3_BAIXA")
            render_card_clickable(c4, "4_OUTRAS")
        
        # Mostrar filtro ativo e botão limpar
        if st.session_state.filtro_selecionado:
            col_info, col_limpar = st.columns([3, 1])
            with col_info:
                st.info(f"✅ Filtrado por: **{map_btn[st.session_state.filtro_selecionado]['label']}**")
            with col_limpar:
                if st.button("Limpar", key="limpar_filtro"):
                    st.session_state.filtro_selecionado = None
                    st.rerun()
        
        st.markdown("---")

        # --- 3. LISTA DETALHADA (ABAIXO) ---
        st.subheader("📋 Lista Detalhada")
        cols = ['Previsão', 'OP', 'Pedido', 'Código Item', 'Cliente', 'Descrição do Item', 'Quantidade', 'Transportadora', 'Prazo Entrega (dias)']
        cols_validas = [c for c in cols if c in df_kpi.columns]
        # incluir quaisquer outras colunas vindas do cronograma (por exemplo as
        # colunas R–AB que hoje contêm checkbox) para que elas também sejam
        # exibidas na lista detalhada.
        extra = [c for c in df_kpi.columns if c not in cols_validas]
        df_display = df_kpi[cols_validas + extra].copy()
        if 'Previsão' in df_display.columns:
            df_display['Previsão'] = pd.to_datetime(df_display['Previsão']).dt.strftime('%d-%m-%Y')
        
        # Renomear "Código Item" para "Código"
        if 'Código Item' in df_display.columns:
            df_display = df_display.rename(columns={'Código Item': 'Código'})
        
        # Destacar OPs já inspecionadas em VERDE
        def estilizar_linhas_concluidas(df_estilo, df_hist_check):
            """Destaca em verde as linhas já inspecionadas."""
            # Criar máscara com mesma forma do dataframe
            mascara = pd.DataFrame('', index=df_estilo.index, columns=df_estilo.columns)
            
            if not df_hist_check.empty:
                # OPs já inspecionadas
                ops_conclusas = set(df_hist_check['OP'].astype(str).unique()) if 'OP' in df_hist_check.columns else set()
                
                # Aplicar cor verde às linhas das OPs conclusas
                for idx in df_estilo.index:
                    if 'OP' in df_estilo.columns and str(df_estilo.loc[idx, 'OP']) in ops_conclusas:
                        mascara.iloc[idx] = 'background-color: #d4edda; color: #155724; font-weight: bold;'
            
            return mascara
        
        # Aplicar estilização
        df_styled = df_display.style.apply(lambda x: estilizar_linhas_concluidas(df_display, df_hist), axis=None)
        
        # Barra de progresso ao exibir dados
        progress_bar = st.progress(0)
        for i in range(len(df_display)):
            progress_bar.progress((i + 1) / len(df_display))
        
        st.dataframe(df_styled, hide_index=True, use_container_width=True)

# ==========================================
# PÁGINA: INSPEÇÃO
# ==========================================
elif pagina == "🔍 Inspeção":
    st.title("📝 Apontamento de Qualidade")
    
    col_insp, _ = st.columns([1, 2])
    with col_insp:
        aviso_campo_faltante("inspetor", "Selecione o Inspetor!")
        inspetor_selecionado = st.selectbox("👷 Inspetor:", ["", "Alexandre", "Pedro", "Lilian", "Ygor"])
    st.markdown("---")

    df_insp = df_final.copy()

    if df_insp.empty:
        st.warning("⚠️ Nenhuma peça encontrada. Verifique se há arquivos na pasta CRONOGRAMA ou os filtros.")
    else:
        lista_clientes = sorted(df_insp['Cliente'].unique())
        cliente_sel = st.selectbox("1️⃣ Cliente:", lista_clientes)
        df_cli = df_insp[df_insp['Cliente'] == cliente_sel]
        df_cli['Label'] = df_cli['OP'].astype(str) + " - " + df_cli['Descrição do Item'].astype(str)
        op_sel = st.selectbox("2️⃣ Peça/OP:", sorted(df_cli['Label'].unique()))
        
        if op_sel:
            op_num = op_sel.split(" - ")[0]
            
            # Detectar se a OP mudou e resetar checkboxes de classificação
            if st.session_state.op_anterior != op_num:
                resetar_classificacoes()
                st.session_state.op_anterior = op_num
                st.session_state.campos_faltantes = set()  # Limpar avisos ao mudar OP
            
            dados = df_cli[df_cli['OP'].astype(str) == op_num].iloc[0]

            qtd_padrao = 1
            # Procurar pela coluna de quantidade - verificar nome exato primeiro
            for col_nome in ['Qtde Entregar', 'Quantidade', 'QTD', 'Qtd']:
                if col_nome in dados.index and dados[col_nome] != "":
                    try: 
                        qtd_padrao = int(float(str(dados[col_nome]).replace(',', '.')))
                        break
                    except: 
                        pass
            
            # Fallback: procurar qualquer coluna com QTD ou QUANT no nome
            if qtd_padrao == 1:
                for c in dados.index:
                    if "QTD" in c.upper() or "QUANT" in c.upper():
                        try: 
                            qtd_padrao = int(float(str(dados[c]).replace(',', '.')))
                            break
                        except: 
                            pass
            
            c1, c2, c3 = st.columns([2, 1.5, 1])
            c1.info(f"📦 **Item:**\n{dados.get('Descrição do Item', '')}\n\n🔢 **Qtd:** {qtd_padrao}")
            prio = dados['Prioridade_Logistica']
            txt_tr = f"🚚 {dados.get('Transportadora', '')}"
            if "ALTA" in prio: c2.error(f"🔴 **ALTA**\n\n{txt_tr}")
            elif "MEDIA" in prio: c2.warning(f"🟡 **MÉDIA**\n\n{txt_tr}")
            else: c2.info(f"⚪ **NORMAL**\n\n{txt_tr}")
            c3.info(f"📅 **Prev:**\n{dados.get('Previsão', '')}")

            desc = dados.get('Descrição do Item', '').upper()
            is_fuso = "FUSO" in desc
            is_guia = "GUIA" in desc or "TRILHO" in desc
            is_bloco = "BLOCO" in desc or "PATIM" in desc
            med_sobra = extrair_medidas_peca(desc)
            isento_tam = verificar_isencao_tamanho(desc)[1]

            qtd_prod = max(1, qtd_padrao)

            st.markdown("---")
            if isento_tam: st.warning("📏 Isento de medição (Tamanho excessivo)")
            
            # Sistema multi-peça
            eh_multiplo = qtd_prod >= 2
            if eh_multiplo:
                eh_ultima_peca = st.session_state.peca_atual == qtd_prod
                with st.container(border=True):
                    pc1, pc2 = st.columns([3, 2])
                    pc1.markdown(f"📦 **Peça {st.session_state.peca_atual} de {qtd_prod}**")
                    with pc2:
                        btn_proxima_peca = st.button(
                            "➡️ Salvar e Ir para Próxima" if not eh_ultima_peca else "✅ Finalizar",
                            use_container_width=True,
                            type="primary",
                        )
            
            # Data e Hora - individual para cada peça em modo multi-peça
            st.markdown("---")
            ct1, ct2, ct3, ct4 = st.columns(4)
            if eh_multiplo:
                data_prod = ct1.date_input("Data Produção", date.today(), format="DD/MM/YYYY", key=f"data_prod_{st.session_state.peca_atual}")
            else:
                data_prod = ct1.date_input("Data Produção", date.today(), format="DD/MM/YYYY")
            
            # Data de chegada: validar que é hoje ou anterior
            data_cheg_permitida = date.today()
            if eh_multiplo:
                data_cheg = ct2.date_input("Data Chegada", value=date.today(), format="DD/MM/YYYY", key=f"data_cheg_{st.session_state.peca_atual}")
            else:
                data_cheg = ct2.date_input("Data Chegada", value=date.today(), format="DD/MM/YYYY")
            if data_cheg > data_cheg_permitida:
                st.error("❌ Data de chegada não pode ser no futuro!")
                data_cheg = data_cheg_permitida
            
            if eh_multiplo:
                ct3.text_input("Hora", key=f"hora_entrada_{st.session_state.peca_atual}", on_change=formatar_hora_automatica, args=(f"hora_entrada_{st.session_state.peca_atual}",))
            else:
                ct3.text_input("Hora", key="hora_entrada_key", on_change=formatar_hora_automatica, args=("hora_entrada_key",))
            ct4.text_input("Líder", value="Pedro Miguel", disabled=True)
            
            cm1, cm2, cm3 = st.columns([1, 1.5, 1])
            cm1.markdown("#### Medições")
            cm2.radio("Escolha o Relógio:", ["Centesimal", "Milesimal"], horizontal=True, key="tipo_relogio_key", on_change=ajustar_casas_relogio)
            
            # Código de Castanha - com suporte a múltiplas peças
            aviso_campo_faltante("cod_cas", "Preencha o Código Castanha!")
            if eh_multiplo:
                cod_cas_key = f"cod_cas_{st.session_state.peca_atual}"
                cod_cas = cm3.text_input("Cód. Castanha", key=cod_cas_key)
            else:
                cod_cas = cm3.text_input("Cód. Castanha")
            
            placeholder_texto = "Insira o valor"
            
            c_esq, c_dir = st.columns(2)
            with c_esq:
                if eh_multiplo:
                    st.text_input("Empenamento ESQ", placeholder=placeholder_texto, key=f"emp_e_{st.session_state.peca_atual}", on_change=converter_medida, args=(f"emp_e_{st.session_state.peca_atual}",))
                    st.text_input("Batimento ESQ", placeholder=placeholder_texto, key=f"bat_e_{st.session_state.peca_atual}", on_change=converter_medida, args=(f"bat_e_{st.session_state.peca_atual}",))
                else:
                    st.text_input("Empenamento ESQ", placeholder=placeholder_texto, key="emp_e", on_change=converter_medida, args=("emp_e",))
                    st.text_input("Batimento ESQ", placeholder=placeholder_texto, key="bat_e", on_change=converter_medida, args=("bat_e",))
            with c_dir:
                if eh_multiplo:
                    st.text_input("Batimento DIR", placeholder=placeholder_texto, key=f"bat_d_{st.session_state.peca_atual}", on_change=converter_medida, args=(f"bat_d_{st.session_state.peca_atual}",))
                    st.text_input("Empenamento DIR", placeholder=placeholder_texto, key=f"emp_d_{st.session_state.peca_atual}", on_change=converter_medida, args=(f"emp_d_{st.session_state.peca_atual}",))
                else:
                    st.text_input("Batimento DIR", placeholder=placeholder_texto, key="bat_d", on_change=converter_medida, args=("bat_d",))
                    st.text_input("Empenamento DIR", placeholder=placeholder_texto, key="emp_d", on_change=converter_medida, args=("emp_d",))

            # Validação de tolerância
            if eh_multiplo:
                vals_str = [st.session_state.get(k) for k in [f"emp_e_{st.session_state.peca_atual}", f"bat_e_{st.session_state.peca_atual}", 
                                                           f"bat_d_{st.session_state.peca_atual}", f"emp_d_{st.session_state.peca_atual}"] 
                        if st.session_state.get(k)]
            else:
                vals_str = [st.session_state[k] for k in ['emp_e', 'bat_e', 'bat_d', 'emp_d'] if st.session_state[k]]
            
            vals = []
            for v in vals_str:
                try:
                    vals.append(float(str(v).strip()))
                except:
                    pass
            
            max_val = max(vals) if vals else 0.0
            tol_limit = 0.05 if "Milesimal" in st.session_state.tipo_relogio_key else 0.5
            liberado = False
            if max_val > tol_limit:
                st.error("⚠️ Fora da tolerância!")
                liberado = st.checkbox("Líder liberou?")

            st.markdown("#### Classificação")
            
            # Detectar automaticamente a classificação baseada na classe de precisão
            class_auto = detectar_classe_precisao(dados.get('Descrição do Item', '').upper())
            
            # ⚠️ IMPORTANTE: Atualizar a session_state ANTES de renderizar os checkboxes
            # Isso garante que o Streamlit não sobrescreva com valores antigos
            if st.session_state.get("class_auto_op") != op_num:
                for chave_checkbox in ['fuso_ret', 'fuso_ret_ad', 'cast_ret', 'cast_ret_ad', 'fuso_lam', 'fuso_lam_ad', 'cast_lam', 'cast_lam_ad', 'guia', 'bloco']:
                    # Mapear chave do checkbox para chave do resultado da classificação
                    chave_resultado = 'sel_' + chave_checkbox
                    if chave_resultado in class_auto:
                        st.session_state[chave_checkbox] = class_auto[chave_resultado]
                st.session_state.class_auto_op = op_num
            
            # RETIFICADOS
            st.markdown("**🔧 RETIFICADOS:**")
            rc1, rc2, rc3, rc4 = st.columns(4)
            sel_fuso_ret = rc1.checkbox("Fuso Retificado", value=st.session_state.get("fuso_ret", class_auto["sel_fuso_ret"]), key="fuso_ret")
            sel_fuso_ret_ad = rc2.checkbox("Fuso Retificado Adaptado", value=st.session_state.get("fuso_ret_ad", class_auto["sel_fuso_ret_ad"]), key="fuso_ret_ad")
            sel_cast_ret = rc3.checkbox("Castanha Retificada", value=st.session_state.get("cast_ret", class_auto["sel_cast_ret"]), key="cast_ret")
            sel_cast_ret_ad = rc4.checkbox("Castanha Ret. Adaptada", value=st.session_state.get("cast_ret_ad", class_auto["sel_cast_ret_ad"]), key="cast_ret_ad")
            
            # LAMINADOS
            st.markdown("**⚙️ LAMINADOS:**")
            lc1, lc2, lc3, lc4 = st.columns(4)
            sel_fuso_lam = lc1.checkbox("Fuso Laminado", value=st.session_state.get("fuso_lam", class_auto["sel_fuso_lam"]), key="fuso_lam")
            sel_fuso_lam_ad = lc2.checkbox("Fuso Laminado Adaptado", value=st.session_state.get("fuso_lam_ad", class_auto["sel_fuso_lam_ad"]), key="fuso_lam_ad")
            sel_cast_lam = lc3.checkbox("Castanha Laminada", value=st.session_state.get("cast_lam", class_auto["sel_cast_lam"]), key="cast_lam")
            sel_cast_lam_ad = lc4.checkbox("Castanha Lam. Adaptada", value=st.session_state.get("cast_lam_ad", class_auto["sel_cast_lam_ad"]), key="cast_lam_ad")
            
            # ISOLADOS (Bloco e Guia não podem ser misturados com Fusos/Castanhas)
            st.markdown("**📦 ESPECIAIS:**")
            oc1, oc2 = st.columns([1, 1])
            sel_guia = oc1.checkbox("Guia", value=st.session_state.get("guia", class_auto["sel_guia"]), key="guia")
            sel_bloco = oc2.checkbox("Bloco", value=st.session_state.get("bloco", class_auto["sel_bloco"]), key="bloco")
            
            # Validação: Bloco e Guia não podem ser selecionados com Fusos/Castanhas
            retificados_selecionados = sel_fuso_ret or sel_fuso_ret_ad or sel_cast_ret or sel_cast_ret_ad
            laminados_selecionados = sel_fuso_lam or sel_fuso_lam_ad or sel_cast_lam or sel_cast_lam_ad
            especiais_selecionados = sel_guia or sel_bloco
            
            # VALIDAÇÃO 1: Não pode ter RETIFICADO e LAMINADO simultaneamente
            if retificados_selecionados and laminados_selecionados:
                st.error("❌ Não pode selecionar RETIFICADO e LAMINADO ao mesmo tempo!")
                # Desmarcar laminados (manter retificados que foram selecionados manualmente)
                sel_fuso_lam = sel_fuso_lam_ad = sel_cast_lam = sel_cast_lam_ad = False
            
            # VALIDAÇÃO 2: Não pode ter ESPECIAIS (Guia/Bloco) com RETIFICADO ou LAMINADO
            if especiais_selecionados and (retificados_selecionados or laminados_selecionados):
                st.error("❌ GUIA e BLOCO não podem ser selecionados com RETIFICADO ou LAMINADO!")
                # Desmarcar retificados e laminados (manter especiais que foram selecionados manualmente)
                sel_fuso_ret = sel_fuso_ret_ad = sel_cast_ret = sel_cast_ret_ad = False
                sel_fuso_lam = sel_fuso_lam_ad = sel_cast_lam = sel_cast_lam_ad = False

            st.markdown("---")
            st.markdown("##### ⚠️ Refugo / RNC")
            tipo_ref_limit = st.session_state.get("tipo_ref", "")
            extra_morte = 0
            if "MORTE" in str(tipo_ref_limit):
                extra_morte = st.number_input("Qtd Extra (MORTE)", min_value=0, max_value=10, key="extra_morte")
            max_refugo = int(qtd_prod) + int(extra_morte) if "MORTE" in str(tipo_ref_limit) else int(qtd_prod)
            r1, r2 = st.columns([2, 3])
            qtd_ref = r1.number_input("Qtd Reprovada", min_value=0, max_value=max_refugo, value=0, key="qtd_ref")

            aviso_campo_faltante("tipo_rnc", "Selecione o Tipo RNC (Fuso/Castanha ou Guia/Bloco)!")
            # Tipo da RNC (nao permitir misturar grupos)
            bloqueia_guia_bloco = is_fuso
            bloqueia_fuso_cast = is_guia or is_bloco
            r2.markdown("**Tipo RNC**")
            c1, c2, c3, c4 = r2.columns(4)
            rnc_fuso = c1.checkbox("Fuso", key="rnc_fuso", disabled=bloqueia_fuso_cast)
            rnc_castanha = c2.checkbox("Castanha", key="rnc_castanha", disabled=bloqueia_fuso_cast)
            rnc_guia = c3.checkbox("Guia", key="rnc_guia", disabled=bloqueia_guia_bloco)
            rnc_bloco = c4.checkbox("Bloco", key="rnc_bloco", disabled=bloqueia_guia_bloco)

            grupo_fc = rnc_fuso or rnc_castanha
            grupo_gb = rnc_guia or rnc_bloco
            if grupo_fc and grupo_gb:
                st.error("❌ Não pode selecionar Fuso/Castanha junto com Guia/Bloco!")
            
            tipo_ref = None; maq_ref = ""; oper_ref = ""; obs_insp = ""; obs_colab = ""; sobra = 0
            analise = False
            motivos = {}
            causas = {}
            
            if qtd_ref > 0:
                st.error("🔴 Detalhes do Refugo")
                cr1, cr2, cr3 = st.columns([2, 1, 1])
                aviso_campo_faltante("tipo_ref", "Selecione o Tipo de Refugo!")
                tipo_ref = cr1.radio("Tipo", ["RETRABALHO", "MORTE COM SOBRA", "MORTE SEM SOBRA"], horizontal=True, key="tipo_ref")
                analise = cr1.checkbox("Em Análise", key="analise_ref")
                aviso_campo_faltante("maq_ref", "Selecione a Máquina!")
                maq_ref = cr2.selectbox("Máquina", LISTA_MAQUINAS, key="maq_ref")
                aviso_campo_faltante("oper_ref", "Selecione o Torneiro!")
                oper_ref = cr3.selectbox("Torneiro", LISTA_TORNEIROS, key="oper_ref")
                
                if "MORTE" in str(tipo_ref):
                    st.info(f"Sobra sugerida: {med_sobra}mm")
                    sobra = st.number_input("Medida Real Sobra (mm)", value=med_sobra, key="sobra_ref")
                
                aviso_campo_faltante("obs_insp", "Preencha a Avaria (Inspetor)!")
                obs_insp = st.text_area("Avaria (Inspetor):", placeholder="O que foi encontrado...", key="obs_insp")
                aviso_campo_faltante("obs_colab", "Preencha a Causa (Colaborador)!")
                obs_colab = st.text_area("Causa (Colaborador):", placeholder="Justificativa...", key="obs_colab")

                # MOSTRAR 2 QUADRANTES DE REFUGO
                st.markdown("---")
                st.markdown("### 📊 Análise de Refugo")
                
                quad1, quad3 = st.columns(2)
                
                # 1º QUADRANTE: Motivos (R-U da aba Lançamentos, removendo as 3 últimas)
                with quad1:
                    st.markdown("**🎯 1º Quadrante - Motivos**")
                    try:
                        arquivo_refugo = '3.1_DASH_MENSAL_01_26.xlsx'
                        if os.path.exists(arquivo_refugo):
                            df_lanc = pd.read_excel(arquivo_refugo, sheet_name='Lançamentos', nrows=0)
                            # Colunas 17-24 (R-U): Usinagem, Inspeção, Desenho, Programação CNC, 
                            # Produção, Comercial, PCP, RETRABALHO OUTROS DP
                            # Removendo 25-27 (Retrabalho, Morta outros, Morta usin.)
                            cols_motivos = df_lanc.columns[17:25].tolist() if len(df_lanc.columns) >= 25 else []
                            for col in cols_motivos:
                                motivos[col] = st.checkbox(col, key=f"mot_{col}")
                        else:
                            st.write("Arquivo 3.1_DASH_MENSAL não encontrado")
                    except Exception as e:
                        st.write(f"Erro: {str(e)}")
                
                # 3º QUADRANTE: Colunas X-AF da aba PRODUTOS DE REFUGO (Causas)
                with quad3:
                    st.markdown("**🔍 3º Quadrante - Causas**")
                    try:
                        if os.path.exists('3.1_DASH_MENSAL_01_26.xlsx'):
                            df_refugo = pd.read_excel('3.1_DASH_MENSAL_01_26.xlsx', 
                                                     sheet_name='PRODUTOS DE REFUGO', header=1, nrows=0)
                            if len(df_refugo.columns) >= 32:
                                cols_x_af = df_refugo.columns[23:32]
                                for col in cols_x_af:
                                    causas[col] = st.checkbox(col, key=f"xaf_{col}")
                        else:
                            st.write("Arquivo não encontrado")
                    except Exception as e:
                        st.write(f"Erro: {str(e)}")

            # Botão para peça única (single-peça) - colocado ao lado do header
            if not eh_multiplo:
                pc1, pc2 = st.columns([3, 2])
                pc1.write("")
                btn_proximo = pc2.button("✅ SALVAR", use_container_width=True, type="primary")
            else:
                btn_proximo = btn_proxima_peca
            
            if btn_proximo:
                erros = []
                avisos = []
                # Resetar campos faltantes
                st.session_state.campos_faltantes = set()
                
                if eh_multiplo:
                    hora_entrada = st.session_state.get(f"hora_entrada_{st.session_state.peca_atual}", "")
                else:
                    hora_entrada = st.session_state.get("hora_entrada_key", "")
                
                # Validação 1: Inspetor obrigatório
                if not inspetor_selecionado: 
                    st.session_state.campos_faltantes.add("inspetor")
                    erros.append("Selecione o Inspetor!")
                
                # Validação 2: Castanha - verificar se é fuso sem castanha (padrão R##-)
                padroes_fuso_sem_castanha = [r'\b[2485]R-', r'\bR\d+-', r'\bL\d+-']
                eh_fuso_sem_castanha = any(re.search(p, desc) for p in padroes_fuso_sem_castanha)
                
                if not cod_cas and not eh_fuso_sem_castanha:
                    st.session_state.campos_faltantes.add("cod_cas")
                    erros.append("Código Castanha obrigatório!")

                # Validações específicas quando há RNC
                if qtd_ref > 0:
                    if not tipo_ref:
                        st.session_state.campos_faltantes.add("tipo_ref")
                        erros.append("Selecione o Tipo de Refugo!")
                    if not maq_ref:
                        st.session_state.campos_faltantes.add("maq_ref")
                        erros.append("Selecione a Máquina!")
                    if not oper_ref:
                        st.session_state.campos_faltantes.add("oper_ref")
                        erros.append("Selecione o Torneiro!")
                    if not (rnc_fuso or rnc_castanha or rnc_guia or rnc_bloco):
                        st.session_state.campos_faltantes.add("tipo_rnc")
                        erros.append("Selecione o Tipo RNC (Fuso/Castanha ou Guia/Bloco)!")
                    if (rnc_fuso or rnc_castanha) and (rnc_guia or rnc_bloco):
                        erros.append("Não pode misturar Fuso/Castanha com Guia/Bloco!")
                    if not str(obs_insp).strip():
                        st.session_state.campos_faltantes.add("obs_insp")
                        erros.append("Preencha a Avaria (Inspetor)!")
                    if not str(obs_colab).strip():
                        st.session_state.campos_faltantes.add("obs_colab")
                        erros.append("Preencha a Causa (Colaborador)!")
                    if not str(hora_entrada).strip():
                        st.session_state.campos_faltantes.add("hora_entrada")
                        erros.append("Preencha a Hora de Entrada!")
                
                # Validação 3: Se for FUSO, precisa ter medição em ESQ ou DIR (ou ambos)
                fuso_selecionado = sel_fuso_ret or sel_fuso_ret_ad or sel_fuso_lam or sel_fuso_lam_ad
                if fuso_selecionado and qtd_ref <= 0:
                    if eh_multiplo:
                        tem_esq = bool(st.session_state.get(f"emp_e_{st.session_state.peca_atual}")) or \
                                  bool(st.session_state.get(f"bat_e_{st.session_state.peca_atual}"))
                        tem_dir = bool(st.session_state.get(f"bat_d_{st.session_state.peca_atual}")) or \
                                  bool(st.session_state.get(f"emp_d_{st.session_state.peca_atual}"))
                    else:
                        tem_esq = bool(st.session_state.get("emp_e")) or bool(st.session_state.get("bat_e"))
                        tem_dir = bool(st.session_state.get("bat_d")) or bool(st.session_state.get("emp_d"))
                    
                    if not (tem_esq or tem_dir):
                        erros.append("Fuso requer medição de pelo menos um lado (ESQ ou DIR)!")
                
                # Validação 4: Tolerância
                if max_val > tol_limit and not liberado: 
                    erros.append("Medida fora sem liberação!")
                
                # Alerta: Castanha repetida (verificar histórico)
                if cod_cas and not eh_fuso_sem_castanha:
                    df_hist_temp = carregar_dados_historico()
                    if not df_hist_temp.empty:
                        # Buscar castanhas já usadas na mesma OP (se a coluna existir)
                        if 'Código Castanha' in df_hist_temp.columns:
                            historico_op = df_hist_temp[df_hist_temp['OP'].astype(str) == str(op_num)]
                            if len(historico_op) > 0 and cod_cas in historico_op['Código Castanha'].values:
                                avisos.append(f"⚠️ Castanha '{cod_cas}' já foi usada nesta OP!")
                
                if erros:
                    for e in erros: st.error(e)
                    if avisos:
                        for av in avisos: st.warning(av)
                    st.rerun()  # Recarregar para mostrar avisos piscantes
                else:
                    if avisos:
                        for av in avisos: st.warning(av)

                    descricao_item = dados.get('Descrição do Item', '')
                    codigo_item = dados.get('Codigo', dados.get('Código Item', dados.get('Código', dados.get('Item', ''))))
                    pedido = dados.get('Pedido', '')
                    transportadora = dados.get('Transportadora', '')
                    hora_saida = datetime.now().strftime("%H:%M")
                    data_producao = data_prod.strftime("%d/%m/%Y") if data_prod else ""
                    data_chegada = data_cheg.strftime("%d/%m/%Y") if data_cheg else ""
                    aprovado = max(qtd_prod - qtd_ref, 0)
                    reprovado = qtd_ref

                    if eh_multiplo:
                        bat_esq = st.session_state.get(f"bat_e_{st.session_state.peca_atual}") or ""
                        bat_dir = st.session_state.get(f"bat_d_{st.session_state.peca_atual}") or ""
                        emp_esq = st.session_state.get(f"emp_e_{st.session_state.peca_atual}") or ""
                        emp_dir = st.session_state.get(f"emp_d_{st.session_state.peca_atual}") or ""
                    else:
                        bat_esq = st.session_state.get("bat_e") or ""
                        bat_dir = st.session_state.get("bat_d") or ""
                        emp_esq = st.session_state.get("emp_e") or ""
                        emp_dir = st.session_state.get("emp_d") or ""

                    mapa_causas = {
                        "Medida não conforme o projeto": "Medida não conforme",
                        "Usinagem não conforme o projeto": "Usinagem não conforme",
                        "Acabamento Ruim": "Acabamento Ruim",
                        "Peça fora de concentricidade": "Concentricidade",
                        "Peça craterizada": "Craterizada",
                        "Estética (aparência)": "Estética",
                        "Rebarba": "Rebarba",
                        "Faltou usinar a chaveta": "Faltou Chaveta",
                        "Desenho Errado": "Desenho Errado",
                    }
                    causas_db = {mapa_causas.get(k, k): v for k, v in causas.items()}
                    marca = lambda v: "X" if v else ""

                    novo = {
                        'Data Registro': date.today().strftime("%d/%m/%Y"),
                        'Hora Saída': hora_saida,
                        'Inspetor': inspetor_selecionado,
                        'Inspetor Responsável': "Pedro Miguel",
                        'Data Chegada': data_chegada,
                        'Hora Entrada': hora_entrada or "",
                        'Data Produção': data_producao,
                        'OP': op_num,
                        'Cliente': dados['Cliente'],
                        'Descrição': descricao_item,
                        'Transportadora': transportadora,
                        'Pedido': pedido,
                        'Codigo': codigo_item,
                        'Qtd Total': qtd_prod,
                        'Peças produzidas': qtd_prod,
                        'Aprovado': aprovado,
                        'Reprovado': reprovado,
                        'Tipo Refugo': tipo_ref if qtd_ref > 0 else "",
                        'Sobra Medida 1': sobra if qtd_ref > 0 and sobra else "",
                        'Sobra Medida 2': "",
                        'Maquina': maq_ref if qtd_ref > 0 else "",
                        'Operador': oper_ref if qtd_ref > 0 else "",
                        'Cód. Castanha': cod_cas or "",
                        'Bat. Esq': bat_esq,
                        'Bat. Dir': bat_dir,
                        'Emp. Esq': emp_esq,
                        'Emp. Dir': emp_dir,
                        'Obs': f"{obs_insp} | {obs_colab}" if qtd_ref > 0 else "",
                        'Status': "Finalizado",
                        'Aprovação Especial': "SIM" if liberado else "NAO",
                        'Fuso Retificado': marca(sel_fuso_ret),
                        'Fuso Retificado Adaptado': marca(sel_fuso_ret_ad),
                        'Castanha Retificada Adaptada': marca(sel_cast_ret_ad),
                        'Fuso Laminado': marca(sel_fuso_lam),
                        'Fuso Laminado PRECISÃO': "",
                        'Fuso Laminado Adaptado': marca(sel_fuso_lam_ad),
                        'Castanha Laminada Adaptada': marca(sel_cast_lam_ad),
                        'Guia': marca(sel_guia),
                        'Bloco': marca(sel_bloco),
                        'Usinagem': marca(motivos.get('Usinagem', False)),
                        'Inspeção': marca(motivos.get('Inspeção', False)),
                        'Desenho': marca(motivos.get('Desenho', False)),
                        'Programação CNC': marca(motivos.get('Programação CNC', False)),
                        'Produção': marca(motivos.get('Produção', False)),
                        'Gerar op': marca(motivos.get('Gerar op - instrução para corte errado', False)),
                        'PCP': marca(motivos.get('PCP', False)),
                        'Medida não conforme': marca(causas_db.get('Medida não conforme', False)),
                        'Usinagem não conforme': marca(causas_db.get('Usinagem não conforme', False)),
                        'Acabamento Ruim': marca(causas_db.get('Acabamento Ruim', False)),
                        'Concentricidade': marca(causas_db.get('Concentricidade', False)),
                        'Craterizada': marca(causas_db.get('Craterizada', False)),
                        'Estética': marca(causas_db.get('Estética', False)),
                        'Rebarba': marca(causas_db.get('Rebarba', False)),
                        'Faltou Chaveta': marca(causas_db.get('Faltou Chaveta', False)),
                        'Desenho Errado': marca(causas_db.get('Desenho Errado', False)),
                        'Código Castanha': cod_cas or "",
                        # registrar também motivos para histórico
                        'Motivo_Usinagem': motivos.get('Usinagem', False),
                        'Motivo_Medida': motivos.get('Medida', False),
                        'Motivo_Outros': motivos.get('Outros', False),
                    }
                    
                    # Se for múltiplas peças, armazenar dados e ir para próxima
                    if eh_multiplo:
                        st.session_state.pecas_inspecionadas[st.session_state.peca_atual] = {
                            'cod_castanha': cod_cas,
                            'emp_e': st.session_state.get(f"emp_e_{st.session_state.peca_atual}"),
                            'bat_e': st.session_state.get(f"bat_e_{st.session_state.peca_atual}"),
                            'bat_d': st.session_state.get(f"bat_d_{st.session_state.peca_atual}"),
                            'emp_d': st.session_state.get(f"emp_d_{st.session_state.peca_atual}"),
                            'data_prod': data_producao,
                            'data_cheg': data_chegada,
                            'hora_entrada': hora_entrada,
                        }
                        
                        if eh_ultima_peca:
                            # Última peça - salvar no banco de dados
                            df_novo = pd.DataFrame([novo])
                            try:
                                salvar_no_banco(df_novo, nao_conforme=(qtd_ref > 0))
                            except Exception as e:
                                st.error(f"Erro ao salvar no histórico: {e}")
                            
                            # Mostrar summary
                            st.success(f"✅ 🎉 **Inspeção concluída!** Todas as {qtd_prod} peças foram inspecionadas com sucesso!")
                            st.write("**Resumo das peças inspecionadas:**")
                            for num_peca, dados_peca in st.session_state.pecas_inspecionadas.items():
                                st.write(f"  • Peça {num_peca}: Cód. Castanha = {dados_peca['cod_castanha']} | Data Prod: {dados_peca['data_prod']} | Hora: {dados_peca['hora_entrada']}")
                            st.session_state.peca_atual = 1
                            st.session_state.pecas_inspecionadas = {}
                            st.cache_data.clear()  # Limpar cache para atualizar lista no Home com verde
                            st.rerun()
                        else:
                            # Ir para próxima peça
                            st.session_state.peca_atual += 1
                            # limpar dados de refugo para a próxima peça
                            for k in [
                                "qtd_ref", "tipo_ref", "analise_ref", "maq_ref", "oper_ref",
                                "sobra_ref", "obs_insp", "obs_colab", "extra_morte",
                                "rnc_fuso", "rnc_castanha", "rnc_guia", "rnc_bloco"
                            ]:
                                if k in st.session_state:
                                    del st.session_state[k]
                            for k in list(st.session_state.keys()):
                                if k.startswith("mot_") or k.startswith("xaf_"):
                                    del st.session_state[k]
                            st.session_state.campos_faltantes = set()  # Limpar avisos ao avançar peça
                            st.info(f"✅ Peça {st.session_state.peca_atual - 1} salva! Carregando peça {st.session_state.peca_atual}...")
                            st.rerun()
                    else:
                        # Peça única - salvar normalmente
                        df_novo = pd.DataFrame([novo])
                        try:
                            salvar_no_banco(df_novo, nao_conforme=(qtd_ref > 0))
                            
                            msg = "Salvo com sucesso!"
                            if qtd_ref > 0:
                                pack = {
                                    'OP': op_num, 'Cliente': dados['Cliente'], 'Descricao': desc,
                                    'Qtd_Total': qtd_prod, 'Qtd_Reprovada': qtd_ref,
                                    'Pedido': dados.get('Pedido', ''), 'Maquina': maq_ref, 'Operador': oper_ref,
                                    'Tipo_Refugo': tipo_ref, 'Analise': analise,
                                    # manter apenas texto livre na descrição; os motivos viram
                                    # flags booleanas tratadas por _campos_para_rnc
                                    'Descricao_Ocorrido': obs_insp,
                                    'Obs_Inspetor': obs_insp, 'Obs_Colaborador': obs_colab,
                                    'Sobra1': sobra, 'Inspetor': inspetor_selecionado,
                                    'Motivo_Usinagem': motivos.get('Usinagem', False),
                                    'Motivo_Medida': motivos.get('Medida', False),
                                    'Motivo_Outros': motivos.get('Outros', False),
                                }
                                nome_rnc = gerar_nome_arquivo_rnc(pack['Pedido'], pack['Cliente'], pack['Descricao'])
                                if preencher_modelo_rnc_existente(pack, os.path.basename(nome_rnc)):
                                    msg += f" RNC gerada: {os.path.basename(nome_rnc)}"
                            
                            st.success(msg)
                            st.cache_data.clear()  # Limpar cache para atualizar lista no Home com verde
                            st.rerun()
                        except Exception as e: st.error(f"Erro ao salvar: {e}")

# ==========================================
# PÁGINA: ANÁLISE DE REFUGO (4 QUADRANTES)
# ==========================================
elif pagina == "♻️ Análise Refugo":
    st.title("♻️ Análise de Ocorrências de Refugo")
    st.markdown("Análise detalhada dos refugos com indicadores e dados de cada ocorrência")
    st.markdown("---")
    
    # Carregar indicadores das abas CONFIGURAÇÃO e Indicadores Usinagem
    try:
        arquivo_dash = '3.1_DASH_MENSAL_01_26.xlsx'
        if os.path.exists(arquivo_dash):
            # ABA CONFIGURAÇÃO - linha 1 tem os totais
            df_config = pd.read_excel(arquivo_dash, sheet_name='CONFIGURAÇÃO ', header=0)
            
            # ABA Indicadores Usinagem
            df_indic = pd.read_excel(arquivo_dash, sheet_name='Indicadores Usinagem')
            
            # Mostrar indicadores principais
            st.markdown("### 📊 Indicadores do Mês")
            
            col_ind1, col_ind2, col_ind3, col_ind4, col_ind5 = st.columns(5)
            
            with col_ind1:
                aprovado = df_config.iloc[1, 1] if len(df_config) > 1 else 0
                st.metric("✅ Aprovado", f"{aprovado}")
            
            with col_ind2:
                reprovado = df_config.iloc[1, 2] if len(df_config) > 1 else 0
                st.metric("❌ Reprovado", f"{reprovado}")
            
            with col_ind3:
                retrabalho_total = df_config.iloc[1, 17] if len(df_config) > 1 and len(df_config.columns) > 17 else 0
                st.metric("🔧 Retrabalho", f"{retrabalho_total}")
            
            with col_ind4:
                morta_total = df_config.iloc[1, 16] if len(df_config) > 1 and len(df_config.columns) > 16 else 0
                st.metric("💀 Morta", f"{morta_total}")
            
            with col_ind5:
                pecas_produzidas = df_indic.iloc[3, 1] if len(df_indic) > 3 else 0
                st.metric("📦 Produzidas", f"{pecas_produzidas}")
            
            st.markdown("---")
            
            # Detalhamento por departamento
            st.markdown("### 📋 Refugo por Departamento")
            col_dep1, col_dep2, col_dep3, col_dep4 = st.columns(4)
            
            with col_dep1:
                usinagem = df_config.iloc[1, 3] if len(df_config) > 1 and len(df_config.columns) > 3 else 0
                inspecao = df_config.iloc[1, 4] if len(df_config) > 1 and len(df_config.columns) > 4 else 0
                st.write(f"**⚙️ Usinagem:** {usinagem}")
                st.write(f"**🔍 Inspeção:** {inspecao}")
            
            with col_dep2:
                desenho = df_config.iloc[1, 5] if len(df_config) > 1 and len(df_config.columns) > 5 else 0
                prog_cnc = df_config.iloc[1, 6] if len(df_config) > 1 and len(df_config.columns) > 6 else 0
                st.write(f"**📐 Desenho:** {desenho}")
                st.write(f"**💻 Prog. CNC:** {prog_cnc}")
            
            with col_dep3:
                producao = df_config.iloc[1, 7] if len(df_config) > 1 and len(df_config.columns) > 7 else 0
                comercial = df_config.iloc[1, 8] if len(df_config) > 1 and len(df_config.columns) > 8 else 0
                st.write(f"**🏭 Produção:** {producao}")
                st.write(f"**💼 Comercial:** {comercial}")
            
            with col_dep4:
                pcp = df_config.iloc[1, 9] if len(df_config) > 1 and len(df_config.columns) > 9 else 0
                st.write(f"**📊 PCP:** {pcp}")
            
            st.markdown("---")
            
    except Exception as e:
        st.warning(f"⚠️ Não foi possível carregar indicadores: {str(e)}")
    
    # Carregar dados de refugo
    df_refugo = carregar_dados_refugo()
    
    if df_refugo is None or df_refugo.empty:
        st.warning("⚠️ Dados de refugo não encontrados. Verifique se o arquivo está carregado.")
    else:
        # Filtros
        col_filtro1, col_filtro2 = st.columns(2)
        
        with col_filtro1:
            fatura = st.text_input("🔍 Filtrar por OP ou Pedido:", placeholder="ex: 04693101001")
        
        with col_filtro2:
            clientes_list = [c for c in df_refugo.get('Cliente', []) if c] if 'Cliente' in df_refugo.columns else []
            cliente_sel = st.selectbox("👥 Cliente:", ["Todos"] + sorted(set(clientes_list)))
        
        # Aplicar filtros
        df_filtrado = df_refugo.copy()
        
        if fatura:
            mask = False
            if 'OP' in df_filtrado.columns:
                mask = mask | df_filtrado['OP'].astype(str).str.contains(fatura, na=False)
            if 'Pedido' in df_filtrado.columns:
                mask = mask | df_filtrado['Pedido'].astype(str).str.contains(fatura, na=False)
            df_filtrado = df_filtrado[mask]
        
        if cliente_sel != "Todos" and 'Cliente' in df_filtrado.columns:
            df_filtrado = df_filtrado[df_filtrado['Cliente'] == cliente_sel]
        
        if df_filtrado.empty:
            st.info("📭 Nenhum resultado encontrado.")
        else:
            st.subheader(f"📊 {len(df_filtrado)} Ocorrência(s) de Refugo")
            
            # Iterar sobre cada ocorrência de refugo
            for idx, row in df_filtrado.iterrows():
                with st.container(border=True):
                    # Header
                    col_h1, col_h2, col_h3 = st.columns([1.5, 1.5, 1])
                    with col_h1:
                        op = row.get('OP', 'N/A')
                        st.markdown(f"### 📋 OP: **{op}**")
                    with col_h2:
                        cliente = row.get('Cliente', 'N/A')
                        st.markdown(f"👥 **{cliente}**")
                    with col_h3:
                        data = row.get('Data', 'N/A')
                        st.caption(f"📅 {data}")
                    
                    st.markdown("---")
                    
                    # ============================================================
                    # 4 QUADRANTES PRINCIPAIS
                    # ============================================================
                    quad1, quad2, quad3, quad4 = st.columns(4)
                    
                    # 1º QUADRANTE: MOTIVOS (R-AB da aba Lançamentos via header de PRODUTOS DE REFUGO Q-W)
                    with quad1:
                        st.markdown("### 🎯 **1º Quadrante | Motivos**")
                        motivos_q1 = ['Usinagem', 'Inspeção', 'Desenho', 'Programação CNC', 
                                     'Produção', 'Gerar op - instrução para corte errado', 'PCP']
                        for col in motivos_q1:
                            if col in row.index:
                                val = row[col]
                                status = "✅" if val == "X" else "⬜"
                                st.write(f"{status} {col}")
                    
                    # 2º QUADRANTE: Q-W (Usinagem, Inspeção, etc da aba PRODUTOS)
                    with quad2:
                        st.markdown("### 📋 **2º Quadrante | Análise 1 (Q-W)**")
                        quad2_cols = ['Usinagem', 'Inspeção', 'Desenho', 'Programação CNC', 
                                     'Produção', 'Gerar OP - Instrução para corte errado', 'PCP']
                        for col in quad2_cols:
                            if col in row.index:
                                val = row[col]
                                status = "✅" if val == "X" else "⬜"
                                st.write(f"{status} {col[:25]}")
                    
                    # 3º QUADRANTE: X-AF (Medida, Acabamento, etc)
                    with quad3:
                        st.markdown("### 🔍 **3º Quadrante | Análise 2 (X-AF)**")
                        quad3_cols = ['Medida não conforme o projeto',
                                     'Usinagem não conforme o projeto',
                                     'Acabamento ruim',
                                     'Peça fora de concentricidade',
                                     'Peça craterizada',
                                     'Estética (Aparência)',
                                     'Rebarba',
                                     'Faltou usinar a chaveta',
                                     'Desenho errado']
                        for col in quad3_cols:
                            if col in row.index:
                                val = row[col]
                                status = "✅" if val == "X" else "⬜"
                                label = col.replace('não conforme', 'n/conf').replace('Peça ', 'P. ')
                                st.write(f"{status} {label[:28]}")
                    
                    # 4º QUADRANTE: F-P (Máquina, Fuso, Castanha, etc)
                    with quad4:
                        st.markdown("### 📊 **4º Quadrante | Dados (F-P)**")
                        quad4_cols = [
                            ('Maquina', 'Máquina'),
                            ('Fuso', 'Fuso'),
                            ('Castanha', 'Castanha'),
                            ('Guia', 'Guia'),
                            ('Bloco', 'Bloco'),
                            ('QTD. DA OP.', 'QTD OP'),
                            ('QTD. PEÇAS CHEGOU', 'Peças'),
                            ('QNT. REPROVADO', 'Reprov.'),
                            ('QNT. RETRABALHADO', 'Retrab.'),
                            ('QNT. USINADO NOVAMENTE', 'Usin. Novo'),
                            ('APROVADO', 'Aprovado'),
                        ]
                        for col_name, label in quad4_cols:
                            if col_name in row.index:
                                val = row[col_name]
                                st.write(f"• {label}: {val if val else '—'}")
                    
                    # Observações finais
                    st.markdown("---")
                    obs = row.get('Observação', '')
                    if obs:
                        st.markdown(f"📝 **Observações:**  \n{obs}")




elif pagina == "📊 Indicadores":
    st.title("📊 Indicadores da Qualidade")
    st.markdown("Acompanhamento do desempenho diário e mensal da qualidade")
    st.markdown("---")
    
    arquivo_dash = '3.1_DASH_MENSAL_01_26.xlsx'
    
    if not os.path.exists(arquivo_dash):
        st.error(f"Arquivo não encontrado: {arquivo_dash}")
        st.info("Faça upload do arquivo 3.1_DASH_MENSAL_01_26.xlsx")
    else:
        try:
            st.markdown("### 📦 Peças Inspecionadas")
            with st.spinner("Carregando..."):
                df_lanc = pd.read_excel(arquivo_dash, sheet_name='Lançamentos', header=None)
                valores = df_lanc.iloc[298, 2:10].tolist()
                meses = ['Fuso Retificado', 'Fuso Retificado Adaptado', 'Castanha Retificada Adaptada', 'Fuso Laminado', 'Castanha Laminada Adaptada', 'Fuso Laminado Adaptado', 'Guia', 'Bloco']
                
                cols = st.columns(4)
                for i in range(4):
                    v = int(valores[i]) if pd.notna(valores[i]) else 0
                    cols[i].metric(f"{meses[i]}", f"{v} peças")
                
                cols2 = st.columns(4)
                for i in range(4):
                    v = int(valores[i+4]) if pd.notna(valores[i+4]) else 0
                    cols2[i].metric(f"{meses[i+4]}", f"{v} peças")
                
                total = sum([int(v) if pd.notna(v) else 0 for v in valores])
                st.markdown("---")
                t1, t2, t3 = st.columns(3)
                t1.metric("Total", f"{total} peças")
                t2.metric("Média Mensal", f"{total//8} peças")
                t3.metric("Período", "8 meses")
                
                st.markdown("#### Evolução Mensal")
                df_g = pd.DataFrame({'Mês': meses, 'Peças': [int(v) if pd.notna(v) else 0 for v in valores]})
                st.bar_chart(df_g.set_index('Mês'))
            
            st.markdown("---")
            st.markdown("### ⚙️ Configuração")
            try:
                df_cfg = pd.read_excel(arquivo_dash, sheet_name='CONFIGURAÇÃO ', header=None)
                st.dataframe(df_cfg.head(30), use_container_width=True)
            except Exception as e:
                st.warning(f"Erro: {e}")
            
            st.markdown("---")
            st.markdown("### 🔧 Indicadores Usinagem")
            try:
                df_usin = pd.read_excel(arquivo_dash, sheet_name='Indicadores Usinagem', header=None)
                st.dataframe(df_usin.head(30), use_container_width=True)
            except Exception as e:
                st.warning(f"Erro: {e}")
        except Exception as e:
            st.error(f"Erro ao carregar: {e}")
            st.exception(e)

elif pagina == "�📦 Pré Carga":
    st.title("📦 Pré Carga")
    st.info("Módulo em desenvolvimento...")