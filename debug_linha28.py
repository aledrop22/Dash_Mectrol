import pandas as pd
from openpyxl import load_workbook

# Usar openpyxl para ver exatamente o que está na planilha
wb = load_workbook('3.1_DASH_MENSAL_01_26.xlsx')
ws = wb['Indicadores Usinagem']

print("=== VERIFICAÇÃO DA LINHA 28 ===\n")
print(f"B28: {ws['B28'].value}")
print(f"C28: {ws['C28'].value}")
print(f"D28: {ws['D28'].value}")
print(f"E28: {ws['E28'].value}")
print(f"F28: {ws['F28'].value}")

print("\n=== CONTEXTO - LINHAS 1 A 30 (COLUNAS A-E) ===\n")
for row in range(1, 31):
    a = ws.cell(row, 1).value
    b = ws.cell(row, 2).value
    c = ws.cell(row, 3).value
    d = ws.cell(row, 4).value
    e = ws.cell(row, 5).value
    print(f"Linha {row:2d}: A={str(a)[:15]:15s} B={str(b)[:20]:20s} C={str(c)[:20]:20s} D={str(d)[:20]:20s} E={str(e)[:20]:20s}")

print("\n=== VALORES NUMÉRICOS NA LINHA 28 ===")
print(f"B28 (índice 1): {ws.cell(28, 2).value if ws.cell(28, 2).value else 'Vazio'}")
print(f"C28 (índice 2): {ws.cell(28, 3).value if ws.cell(28, 3).value else 'Vazio'}")
print(f"D28 (índice 3): {ws.cell(28, 4).value if ws.cell(28, 4).value else 'Vazio'}")
